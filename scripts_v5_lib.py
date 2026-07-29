# -*- coding: utf-8 -*-
"""
scripts_v5_lib.py — Refactored MSK inter-rater analysis (v5)

Lessons from v4 kept:
  • Rater identity = (sessionUser, sessionUserEmail), not filename
  • true_protocol from groupId (protocolId alone cannot split sulcus-angle vs 3cm)
  • Import coverage diagnostic (counts per rater × patient × protocol)
  • Pairwise redo = ≥10% relative gap AND |z|≥1.96 vs pair Bland–Altman
  • Line endpoint standardization + sulcus condyle anatomical relabeling

v5 changes:
  • All exports go under analysis_outputs_v5/ (never overwrite live redo_tracker.xlsx)
  • Dedup key includes true_protocol (fixes sulcus drop when protocolId collides)
  • JMP long export is the primary data export (ML ground-truth exports removed)
  • Consensus / Difficult / Calibration redo method removed
  • One library + thin notebook instead of 40+ ad-hoc cells
"""

from __future__ import annotations

import json
import os
import subprocess
import sys
from itertools import combinations
from pathlib import Path
from typing import Iterable, Optional

import numpy as np
import pandas as pd

try:
    from IPython.display import display  # type: ignore
except ImportError:
    def display(obj):
        print(obj)


# =============================================================================
# Config
# =============================================================================

LOCAL_PROJECT_PATH = None  # or paste a Finder path
PROJECT_FOLDER_NAME = "Current Knee MRI Project Folder"
CSV_SUBFOLDER_NAME = "csv exports"

# ALL v5 writes land here — never the project root live tracker
OUTPUT_DIR_NAME = "analysis_outputs_v5"
REDO_SUBDIR_NAME = "redo_tracker"

# Filenames that must never be written at the project root by v5
PROTECTED_ROOT_NAMES = {
    "redo_tracker.xlsx",
    "redo_queue.csv",
    "redo_queue_details.csv",
    "redo_queue_by_rater.csv",
    "redo_queue_unique_cases.csv",
    "redo_queue_consensus.csv",
    "redo_queue_consensus_details.csv",
    "redo_queue_difficult.csv",
    "redo_queue_calibration.csv",
    "redo_queue_calibration_summary.csv",
    "redo_queue_consensus_outcomes.csv",
    "jmp_statistical_comparison_long.csv",  # live JMP stays until you promote v5
    "ml_ready_annotations.csv",
    "standardized_landmarks.csv",
}

# Default exclusions (training / incomplete annotators). Override per-call.
EXCLUDE_RATERS = ["Zachary Liu", "Raaga Ramesh", "Justin Lin"]

# Optional whitelist for ICC / JMP / redo. None = all raters minus EXCLUDE_RATERS.
# Example: ICC_INCLUDE_RATERS = ["Danish", "Madalyn", "Miyaz", "Parker", "Roy"]
ICC_INCLUDE_RATERS = None  # type: Optional[list]

# v4 notebook collapsed plain sulcus-angle into sulcus-angle-3cm before ICC.
# That produced the historical ~0.70 sulcus ICC. Keep as default for continuity;
# set False to analyze the two sulcus protocols separately.
COLLAPSE_SULCUS_FOR_ICC = True

PAIR_PCT_MIN = 0.10
BA_Z_MIN = 1.96
REDO_MIN_PAIR_N = 5
STATUS_DEFAULT = "Pending"

PROTOCOL_PCT_FLOOR = {
    "tt-tg": 10.0,
    "patellar-tilt": 5.0,
    "sulcus-angle": 100.0,
    "sulcus-angle-3cm": 100.0,
    "insall-salvati": 0.8,
    "caton-deschamps": 0.8,
}

PROTOCOL_META = {
    "tt-tg": {"unit": "mm", "label": "TT-TG distance"},
    "insall-salvati": {"unit": "ratio", "label": "Insall–Salvati ratio"},
    "patellar-tilt": {"unit": "deg", "label": "Patellar tilt"},
    "sulcus-angle": {"unit": "deg", "label": "Sulcus angle"},
    "sulcus-angle-3cm": {"unit": "deg", "label": "Sulcus angle (3 cm)"},
    "caton-deschamps": {"unit": "ratio", "label": "Caton–Deschamps ratio"},
}

PROTOCOL_DISPLAY = {k: v["label"] for k, v in PROTOCOL_META.items()}

SULCUS_PROTOCOLS = {"sulcus-angle", "sulcus-angle-3cm"}
SULCUS_CONDYLE_STEPS = {"lateral-line", "medial-line"}
SULCUS_GROUP_KEYS = ["patientId", "laterality", "true_protocol", "groupId", "rater_key"]

# One clinical measurement per rater × case × protocol (latest by groupId timestamp).
# Do NOT include sequenceName/plane — those create duplicate case rows after merge
# and break pingouin ICC (duplicate subject IDs).
CASE_KEY = [
    "patientId",
    "laterality",
    "true_protocol",
]


# =============================================================================
# Packages / path helpers
# =============================================================================

def ensure_packages(pkgs: Iterable[tuple[str, str]] | None = None) -> None:
    pkgs = list(pkgs or [
        ("pingouin", "pingouin"),
        ("openpyxl", "openpyxl"),
        ("seaborn", "seaborn"),
        ("scipy", "scipy"),
        ("matplotlib", "matplotlib"),
    ])
    missing = []
    for pkg, mod in pkgs:
        try:
            __import__(mod)
        except ImportError:
            missing.append(pkg)
    if missing:
        print("Installing:", ", ".join(missing))
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", *missing])


def _in_colab() -> bool:
    try:
        import google.colab  # type: ignore  # noqa: F401
        return True
    except ImportError:
        return False


def _candidate_drive_project_paths() -> list[Path]:
    home = Path.home()
    candidates: list[Path] = []
    if LOCAL_PROJECT_PATH:
        candidates.append(Path(LOCAL_PROJECT_PATH).expanduser())
    cloud = home / "Library" / "CloudStorage"
    if cloud.is_dir():
        for entry in sorted(cloud.iterdir()):
            if entry.name.startswith("GoogleDrive"):
                candidates.append(entry / "My Drive" / PROJECT_FOLDER_NAME)
                candidates.append(entry / "MyDrive" / PROJECT_FOLDER_NAME)
                # Shared-drive shortcut layout used by this project
                shortcut_root = entry / ".shortcut-targets-by-id"
                if shortcut_root.is_dir():
                    for sid in shortcut_root.iterdir():
                        candidates.append(sid / PROJECT_FOLDER_NAME)
    for base in (
        home / "Google Drive" / "My Drive",
        home / "Google Drive",
        Path("/Volumes/GoogleDrive/My Drive"),
    ):
        candidates.append(base / PROJECT_FOLDER_NAME)
    try:
        here = Path(__file__).resolve().parent
    except NameError:
        here = Path.cwd()
    candidates.append(here / PROJECT_FOLDER_NAME)
    candidates.append(here)
    candidates.append(here.parent)
    return candidates


def resolve_project_path() -> Path:
    if _in_colab():
        from google.colab import drive  # type: ignore
        drive.mount("/content/drive", force_remount=False)
        for p in (
            Path("/content/drive/MyDrive") / PROJECT_FOLDER_NAME,
            Path("/content/drive/My Drive") / PROJECT_FOLDER_NAME,
        ):
            if (p / CSV_SUBFOLDER_NAME).is_dir():
                return p

    for project in _candidate_drive_project_paths():
        if (project / CSV_SUBFOLDER_NAME).is_dir():
            return project.resolve()

    searched = "\n".join(f"  - {p / CSV_SUBFOLDER_NAME}" for p in _candidate_drive_project_paths())
    raise FileNotFoundError(
        "Could not find the project CSV folder.\n"
        "Install Google Drive for Desktop, or set LOCAL_PROJECT_PATH.\n"
        f"Paths tried:\n{searched}"
    )


def output_dir(project_path: Path | str) -> Path:
    out = Path(project_path) / OUTPUT_DIR_NAME
    out.mkdir(parents=True, exist_ok=True)
    return out


def redo_dir(project_path: Path | str) -> Path:
    d = output_dir(project_path) / REDO_SUBDIR_NAME
    d.mkdir(parents=True, exist_ok=True)
    return d


def safe_write_path(project_path: Path | str, *parts: str) -> Path:
    """Return a path under analysis_outputs_v5/, refusing protected live root files."""
    project_path = Path(project_path).resolve()
    out_root = output_dir(project_path).resolve()
    target = (out_root.joinpath(*parts)).resolve()
    if target != out_root and out_root not in target.parents:
        raise RuntimeError(f"Refusing write outside {OUTPUT_DIR_NAME}: {target}")
    # Extra guard: never write protected names into the project root
    if target.parent.resolve() == project_path and target.name in PROTECTED_ROOT_NAMES:
        raise RuntimeError(
            f"Refusing to overwrite live project file: {target}\n"
            f"v5 writes only under {project_path / OUTPUT_DIR_NAME}"
        )
    target.parent.mkdir(parents=True, exist_ok=True)
    return target


def assert_live_tracker_untouched(project_path: Path | str, before_mtime: float | None = None) -> None:
    live = Path(project_path) / "redo_tracker.xlsx"
    if before_mtime is not None and live.exists():
        after = live.stat().st_mtime
        if after != before_mtime:
            raise RuntimeError(
                f"LIVE redo_tracker.xlsx mtime changed ({before_mtime} → {after}). "
                "Aborting — v5 must never modify the active tracker."
            )


def rater_name(rk: str) -> str:
    return rk.split(" | ")[0]


def _as_str_list(value) -> Optional[list]:
    """Normalize include/exclude args to a list of strings (or None)."""
    if value is None:
        return None
    if isinstance(value, str):
        # Common mistake: pass a single name as a string → iterate characters
        value = [value]
    try:
        out = [str(x).strip() for x in value if str(x).strip()]
    except TypeError as e:
        raise TypeError(
            f"include/exclude must be a list of rater name strings (or None), got {type(value)!r}"
        ) from e
    return out


def active_rater_keys(
    rater_keys,
    exclude=None,
    include=None,
):
    """
    Select raters for ICC / Bland–Altman / redo / JMP.

    - include: if set, keep only keys whose name/key contains any of these substrings
      (case-insensitive). Example: ["Danish", "Madalyn", "Miyaz", "Parker", "Roy"]
    - exclude: drop keys containing any of these substrings (applied after include,
      or alone when include is None). Defaults to EXCLUDE_RATERS.
    """
    keys = [str(rk) for rk in list(rater_keys)]
    if include is None:
        include = ICC_INCLUDE_RATERS
    if exclude is None:
        exclude = EXCLUDE_RATERS

    include = _as_str_list(include)
    exclude = _as_str_list(exclude) or []

    if include:
        keys = [
            rk for rk in keys
            if any(tok.lower() in rk.lower() for tok in include)
        ]
    if exclude:
        keys = [
            rk for rk in keys
            if not any(excl.lower() in rk.lower() for excl in exclude)
        ]
    return keys


def friendly_protocol(p: str) -> str:
    return PROTOCOL_DISPLAY.get(str(p), str(p))


def friendly_side(s: str) -> str:
    s = str(s).strip().lower()
    return {"left": "Left", "right": "Right"}.get(s, str(s).title())


# =============================================================================
# Load + inventory
# =============================================================================

def load_all_exports(project_path: Path | str) -> tuple[pd.DataFrame, dict[str, pd.DataFrame], list[str]]:
    project_path = Path(project_path)
    csv_folder = project_path / CSV_SUBFOLDER_NAME
    if not csv_folder.is_dir():
        raise FileNotFoundError(f"CSV folder not found: {csv_folder}")

    csv_files = sorted(f for f in os.listdir(csv_folder) if f.lower().endswith(".csv"))
    print(f"Found {len(csv_files)} CSV file(s) in {csv_folder}")

    raw_frames = []
    for fname in csv_files:
        fpath = csv_folder / fname
        try:
            fdf = pd.read_csv(fpath, on_bad_lines="warn", engine="python")
        except UnicodeDecodeError:
            try:
                fdf = pd.read_csv(
                    fpath, encoding="cp1252", encoding_errors="replace",
                    on_bad_lines="warn", engine="python",
                )
            except Exception as e:
                print(f"⚠ Skipping unreadable CSV '{fname}': {e}")
                continue
        except Exception as e:
            print(f"⚠ Skipping unreadable CSV '{fname}': {e}")
            continue

        required_cols = {"groupId", "recordType"}
        if not required_cols.issubset(set(fdf.columns)):
            print(
                f"⚠ Skipping '{fname}': missing required columns "
                f"{sorted(required_cols - set(fdf.columns))}"
            )
            continue

        fdf["__source_file"] = fname
        fdf["__file_mtime"] = os.path.getmtime(fpath)
        fdf["__row_index"] = fdf.index
        raw_frames.append(fdf)

    if not raw_frames:
        raise FileNotFoundError(f"No readable CSV files in {csv_folder}")

    all_rows = pd.concat(raw_frames, ignore_index=True)
    # Critical: true_protocol from groupId — NOT protocolId
    all_rows["true_protocol"] = all_rows["groupId"].str.extract(
        r"^([a-z][a-z0-9]*(?:-[a-z0-9]+)*?)-\d+$"
    )

    if "sessionUserEmail" in all_rows.columns and all_rows["sessionUserEmail"].notna().any():
        id_cols = ["sessionUser", "sessionUserEmail"]
    else:
        id_cols = ["__source_file"]

    all_rows["rater_key"] = all_rows[id_cols].astype(str).agg(" | ".join, axis=1)
    dataframes = {key: grp.copy() for key, grp in all_rows.groupby("rater_key")}
    rater_keys = sorted(dataframes.keys())
    print(f"Raters identified: {len(rater_keys)}")
    for rk in rater_keys:
        print(f"  • {rk}")
    return all_rows, dataframes, rater_keys


def coverage_diagnostic(all_rows: pd.DataFrame, rater_keys: list[str] | None = None) -> dict[str, pd.DataFrame]:
    """
    Lesson-learned diagnostic: how many protocol_result measurements each person
    has per patient/side/protocol — the check that revealed sulcus export gaps.
    """
    res = all_rows[all_rows["recordType"] == "protocol_result"].copy()
    if rater_keys is not None:
        res = res[res["rater_key"].isin(rater_keys)]

    print("=== Import coverage diagnostic ===")
    print("\n1) protocol_result counts per rater × true_protocol:")
    raw_counts = (
        res.groupby(["rater_key", "true_protocol"]).size().unstack(fill_value=0)
    )
    raw_counts.index = [rater_name(i) for i in raw_counts.index]
    display(raw_counts.astype(int))

    print("\n2) Case coverage matrix (patient × side × protocol → # active raters):")
    case_cov = (
        res.groupby(["patientId", "laterality", "true_protocol"])["rater_key"]
        .nunique()
        .rename("n_raters")
        .reset_index()
    )
    by_proto = (
        case_cov.groupby(["true_protocol", "n_raters"]).size()
        .unstack(fill_value=0)
        .sort_index()
    )
    display(by_proto)

    print("\n3) Sulcus variants (protocolId collides; true_protocol does not):")
    sulcus = res[res["true_protocol"].isin(SULCUS_PROTOCOLS)]
    if sulcus.empty:
        print("  No sulcus protocol_result rows found.")
        sulcus_table = pd.DataFrame()
    else:
        sulcus_table = (
            sulcus.groupby(["true_protocol", "protocolId"], dropna=False)
            .size()
            .reset_index(name="n")
        )
        display(sulcus_table)
        # Per-case: who has plain sulcus-angle vs 3cm
        pivot = sulcus.pivot_table(
            index=["patientId", "laterality"],
            columns="true_protocol",
            values="rater_key",
            aggfunc="nunique",
            fill_value=0,
        )
        print("\n   Cases with BOTH sulcus-angle and sulcus-angle-3cm present:")
        if set(SULCUS_PROTOCOLS).issubset(pivot.columns):
            both = pivot[(pivot["sulcus-angle"] > 0) & (pivot["sulcus-angle-3cm"] > 0)]
            print(f"   {len(both)} case/side rows")
            display(both.head(10))
        else:
            print("   (only one sulcus variant present in this export set)")

    print("\n4) Steps without matching protocol_result (possible incomplete export):")
    steps = all_rows[all_rows["recordType"] == "step_measurement"]
    step_groups = steps.groupby(["rater_key", "true_protocol"])["groupId"].nunique()
    result_groups = res.groupby(["rater_key", "true_protocol"])["groupId"].nunique()
    gap = (step_groups - result_groups.reindex(step_groups.index).fillna(0)).rename("steps_minus_results")
    gap = gap[gap > 0].sort_values(ascending=False)
    if gap.empty:
        print("  None — every step group has a protocol_result.")
        gap_df = pd.DataFrame()
    else:
        gap_df = gap.reset_index()
        gap_df["rater"] = gap_df["rater_key"].map(rater_name)
        display(gap_df.head(20))

    # Per patient × rater × protocol count (the matrix that caught JMP misalignment)
    print("\n5) Measurement presence matrix (1 = has protocol_result):")
    presence = (
        res.assign(present=1)
        .pivot_table(
            index=["patientId", "laterality", "true_protocol"],
            columns="rater_key",
            values="present",
            aggfunc="max",
            fill_value=0,
        )
    )
    presence.columns = [rater_name(c) for c in presence.columns]
    display(presence.head(15))

    return {
        "raw_counts": raw_counts,
        "case_coverage": case_cov,
        "coverage_by_protocol": by_proto,
        "sulcus_table": sulcus_table,
        "step_result_gaps": gap_df,
        "presence_matrix": presence,
    }


# =============================================================================
# Standardization
# =============================================================================

def parse_points_json(val):
    try:
        pts = json.loads(val) if isinstance(val, str) else val
        return pts if isinstance(pts, list) else []
    except (TypeError, json.JSONDecodeError):
        return []


def serialize_points_json(pts):
    return json.dumps(pts)


def sync_flat_point_columns(row, pts):
    for i, pt in enumerate(pts[:4]):
        for axis in ("x", "y", "z"):
            col = f"p{i}_{axis}"
            if col in row.index:
                row[col] = pt.get(axis, np.nan)
    return row


def line_center_x(pts):
    if len(pts) < 2:
        return np.nan
    return (pts[0]["x"] + pts[1]["x"]) / 2.0


def standardize_line_endpoints(pts):
    if len(pts) < 2:
        return pts, False
    p0, p1 = dict(pts[0]), dict(pts[1])
    dx = abs(p1["x"] - p0["x"])
    dy = abs(p1["y"] - p0["y"])
    swapped = False
    if dx >= dy:
        if p0["x"] > p1["x"]:
            p0, p1 = p1, p0
            swapped = True
    else:
        if p0["y"] > p1["y"]:
            p0, p1 = p1, p0
            swapped = True
    out = [p0, p1] + [dict(p) for p in pts[2:]]
    if swapped and len(out) >= 3:
        out[2]["x"] = out[1]["x"]
        out[2]["y"] = out[1]["y"]
        if "z" in out[1]:
            out[2]["z"] = out[1]["z"]
    return out, swapped


def lateral_is_higher_x(laterality):
    lat = str(laterality).strip().lower()
    if lat == "left":
        return True
    if lat == "right":
        return False
    return None


def apply_endpoint_standardization_to_row(row):
    if row.get("recordType") != "step_measurement":
        return row, False
    if row.get("measurementType") != "distance":
        return row, False
    if pd.isna(row.get("pointCount")) or int(row["pointCount"]) < 2:
        return row, False
    pts = parse_points_json(row.get("stepPointsMmJson"))
    if len(pts) < 2:
        return row, False
    std_pts, swapped = standardize_line_endpoints(pts)
    if not swapped:
        return row, False
    row = row.copy()
    row["stepPointsMmJson"] = serialize_points_json(std_pts)
    if "pointsJson" in row.index and pd.notna(row.get("pointsJson")):
        row["pointsJson"] = serialize_points_json(std_pts)
    row = sync_flat_point_columns(row, std_pts)
    row["endpoints_standardized"] = True
    return row, True


def correct_sulcus_condyle_labels(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "sulcus_label_corrected" not in df.columns:
        df["sulcus_label_corrected"] = False
    if "original_stepId" not in df.columns:
        df["original_stepId"] = df["stepId"]

    labels = {
        "lateral-line": "Line from lateral condyle peak to sulcus",
        "medial-line": "Line from medial condyle peak to sulcus",
    }
    step_mask = (
        (df["recordType"] == "step_measurement")
        & (df["true_protocol"].isin(SULCUS_PROTOCOLS))
        & (df["stepId"].isin(SULCUS_CONDYLE_STEPS))
    )
    if not step_mask.any():
        return df

    for _, grp in df.loc[step_mask].groupby(SULCUS_GROUP_KEYS):
        if len(grp) != 2:
            continue
        idx_a, idx_b = grp.index[0], grp.index[1]
        cx_a = line_center_x(parse_points_json(df.at[idx_a, "stepPointsMmJson"]))
        cx_b = line_center_x(parse_points_json(df.at[idx_b, "stepPointsMmJson"]))
        higher_x_is_lateral = lateral_is_higher_x(df.at[idx_a, "laterality"])
        if higher_x_is_lateral is None:
            continue
        if higher_x_is_lateral:
            target = {
                idx_a: ("lateral-line" if cx_a >= cx_b else "medial-line"),
                idx_b: ("lateral-line" if cx_b > cx_a else "medial-line"),
            }
        else:
            target = {
                idx_a: ("lateral-line" if cx_a <= cx_b else "medial-line"),
                idx_b: ("lateral-line" if cx_b < cx_a else "medial-line"),
            }
        for idx, new_step in target.items():
            if df.at[idx, "stepId"] == new_step:
                continue
            df.at[idx, "original_stepId"] = df.at[idx, "stepId"]
            df.at[idx, "stepId"] = new_step
            if "stepLabel" in df.columns:
                df.at[idx, "stepLabel"] = labels.get(new_step, df.at[idx, "stepLabel"])
            df.at[idx, "sulcus_label_corrected"] = True
    return df


def standardize_rater_dataframe(df: pd.DataFrame, rk: str) -> pd.DataFrame:
    df = df.copy()
    df["rater_key"] = rk
    df["endpoints_standardized"] = False
    df["original_stepId"] = df["stepId"]
    n_endpoints = 0
    for idx in df.index[df["recordType"] == "step_measurement"]:
        new_row, changed = apply_endpoint_standardization_to_row(df.loc[idx])
        if changed:
            df.loc[idx] = new_row
            n_endpoints += 1
    df = correct_sulcus_condyle_labels(df)
    n_sulcus = int(df.get("sulcus_label_corrected", pd.Series(False)).sum())
    print(
        f"  {rater_name(rk)}: {n_endpoints} line(s) endpoint-standardized, "
        f"{n_sulcus} sulcus condyle row(s) relabeled"
    )
    return df


def standardize_all(dataframes: dict[str, pd.DataFrame], rater_keys: list[str]) -> dict[str, pd.DataFrame]:
    print("\nApplying line standardization (endpoints + sulcus anatomical matching)...")
    return {rk: standardize_rater_dataframe(dataframes[rk], rk) for rk in rater_keys}


# =============================================================================
# Comparison table + reliability
# =============================================================================

def extract_results(df: pd.DataFrame, rk: str, collapse_sulcus: Optional[bool] = None) -> pd.DataFrame:
    """Latest protocol_result per clinical case (patient × side × true_protocol)."""
    if collapse_sulcus is None:
        collapse_sulcus = COLLAPSE_SULCUS_FOR_ICC

    result_df = df[df["recordType"] == "protocol_result"].copy()
    if result_df.empty:
        return pd.DataFrame(columns=["patientId", "laterality", "true_protocol", f"value_{rk}", f"unit_{rk}"])

    # Match v4 historical ICC: treat plain sulcus-angle as sulcus-angle-3cm
    if collapse_sulcus and "true_protocol" in result_df.columns:
        result_df["true_protocol"] = result_df["true_protocol"].replace(
            "sulcus-angle", "sulcus-angle-3cm"
        )

    result_df["__group_timestamp"] = pd.to_numeric(
        result_df["groupId"].str.extract(r"-(\d+)$")[0], errors="coerce"
    ).fillna(0)

    # Normalize join keys (avoids int/str patientId merge mismatches)
    result_df["laterality"] = result_df["laterality"].astype(str).str.strip().str.lower()
    result_df["patientId"] = result_df["patientId"].astype(str).str.strip()
    result_df["true_protocol"] = result_df["true_protocol"].astype(str).str.strip()

    subset_cols = [c for c in CASE_KEY if c in result_df.columns]
    sub = result_df.sort_values(["__group_timestamp", "__row_index"], ascending=False)
    before = len(sub)
    sub = sub.drop_duplicates(subset=subset_cols, keep="first")
    n_dup = before - len(sub)
    if n_dup:
        print(f"  {rater_name(rk)}: ignored {n_dup} older duplicate measurement(s)")

    out = sub[["patientId", "laterality", "true_protocol", "resultValue", "resultUnit"]].rename(
        columns={"resultValue": f"value_{rk}", "resultUnit": f"unit_{rk}"}
    )
    out[f"value_{rk}"] = pd.to_numeric(out[f"value_{rk}"], errors="coerce")
    return out


def build_comparison_df(
    dataframes: dict,
    rater_keys,
    min_raters: int = 2,
    collapse_sulcus: Optional[bool] = None,
) -> tuple:
    if collapse_sulcus is None:
        collapse_sulcus = COLLAPSE_SULCUS_FOR_ICC
    print(
        "Extracting protocol_result rows "
        f"(dedupe on patient×side×true_protocol; collapse_sulcus={collapse_sulcus})..."
    )
    processed = [extract_results(dataframes[k], k, collapse_sulcus=collapse_sulcus) for k in rater_keys]
    join_keys = ["patientId", "laterality", "true_protocol"]
    comparison_df = processed[0]
    for nxt in processed[1:]:
        comparison_df = comparison_df.merge(nxt, on=join_keys, how="outer")

    value_cols = [f"value_{k}" for k in rater_keys]
    # Safety: one row per case even if a source still had residual dups
    if comparison_df.duplicated(join_keys).any():
        n_dup = int(comparison_df.duplicated(join_keys).sum())
        print(f"  Collapsing {n_dup} residual duplicate case row(s) via median")
        comparison_df = (
            comparison_df.groupby(join_keys, as_index=False)[value_cols].median()
        )

    comparison_df["n_raters_present"] = comparison_df[value_cols].notna().sum(axis=1)
    comparison_df = comparison_df[comparison_df["n_raters_present"] >= min_raters].reset_index(drop=True)

    print(f"\nMatched records (≥{min_raters} raters): {len(comparison_df)}")
    display(comparison_df.groupby("true_protocol")[value_cols].count())
    return comparison_df, value_cols


def compute_icc_and_pearson(
    comparison_df: pd.DataFrame,
    rater_keys,
    exclude=None,
    include=None,
) -> tuple:
    import pingouin as pg

    keys = active_rater_keys(rater_keys, exclude=exclude, include=include)
    if not keys:
        print("No raters selected for ICC — check include/exclude lists.")
        return pd.DataFrame(), pd.DataFrame()

    value_cols = [f"value_{rk}" for rk in keys if f"value_{rk}" in comparison_df.columns]
    print(f"ICC/Pearson on {len(keys)} active raters: {[rater_name(r) for r in keys]}")
    print(f"  include={_as_str_list(include) if include is not None else ICC_INCLUDE_RATERS}")
    print(f"  exclude={_as_str_list(exclude) if exclude is not None else EXCLUDE_RATERS}")

    icc_results, pearson_results = [], []
    for protocol in sorted(comparison_df["true_protocol"].dropna().unique()):
        prot_df = comparison_df[comparison_df["true_protocol"] == protocol].copy()
        if prot_df.empty:
            continue

        long_df = prot_df.melt(
            id_vars=["patientId", "laterality"],
            value_vars=value_cols,
            var_name="rater",
            value_name="value",
        ).dropna(subset=["value"])
        long_df["value"] = pd.to_numeric(long_df["value"], errors="coerce")
        long_df = long_df.dropna(subset=["value"])
        long_df["subject"] = long_df["patientId"].astype(str) + "_" + long_df["laterality"].astype(str)
        # pingouin requires unique subject×rater rows
        long_df = long_df.groupby(["subject", "rater"], as_index=False)["value"].median()
        counts = long_df["subject"].value_counts()
        long_df = long_df[long_df["subject"].isin(counts[counts >= 2].index)]

        if long_df.empty or long_df["rater"].nunique() < 2 or long_df["subject"].nunique() < 3:
            print(f"  Skipping ICC for {protocol}: insufficient overlapping ratings")
            continue

        try:
            try:
                icc = pg.intraclass_corr(
                    data=long_df, targets="subject", raters="rater",
                    ratings="value", nan_policy="omit",
                )
            except TypeError:
                # Older pingouin has no nan_policy=
                icc = pg.intraclass_corr(
                    data=long_df, targets="subject", raters="rater", ratings="value",
                )
            icc["Protocol"] = protocol
            icc["n_subjects"] = long_df["subject"].nunique()
            icc_results.append(icc)
        except Exception as e:
            print(f"Could not calculate ICC for {protocol}: {e}")

        for ra, rb in combinations(keys, 2):
            va, vb = f"value_{ra}", f"value_{rb}"
            if va not in prot_df.columns or vb not in prot_df.columns:
                continue
            pair = prot_df[[va, vb]].apply(pd.to_numeric, errors="coerce").dropna()
            if len(pair) > 2:
                pearson_results.append({
                    "Protocol": protocol,
                    "Rater A": rater_name(ra),
                    "Rater B": rater_name(rb),
                    "n": len(pair),
                    "Pearson r": round(pair[va].corr(pair[vb]), 4),
                })

    if icc_results:
        icc_df = pd.concat(icc_results)
        icc_summary = icc_df[icc_df["Type"].astype(str).str.contains(r"ICC2|ICC\(A,1\)")].copy()
        ci_col = "CI95" if "CI95" in icc_summary.columns else "CI95%"
        cols = [c for c in ["Protocol", "Type", "ICC", "F", "df1", "df2", "pval", ci_col, "n_subjects"] if c in icc_summary.columns]
        print("\n--- ICC(2,1) per Protocol ---")
        display(icc_summary[cols].round(3))
    else:
        icc_summary = pd.DataFrame()

    pearson_df = pd.DataFrame(pearson_results)
    if len(pearson_df):
        print("\n--- Mean pairwise Pearson r per Protocol ---")
        display(pearson_df.groupby("Protocol")["Pearson r"].mean().to_frame())
    return icc_summary, pearson_df


# =============================================================================
# Bland–Altman / bias summary
# =============================================================================

def bland_altman_and_bias(
    comparison_df: pd.DataFrame,
    rater_keys: list[str],
    exclude: Optional[list[str]] = None,
    make_plots: bool = True,
) -> pd.DataFrame:
    import matplotlib.pyplot as plt

    keys = active_rater_keys(rater_keys, exclude)
    pairs = list(combinations(keys, 2))
    summary_rows = []

    for protocol in sorted(comparison_df["true_protocol"].dropna().unique()):
        prot = comparison_df[comparison_df["true_protocol"] == protocol]
        unit = PROTOCOL_META.get(protocol, {}).get("unit", "")
        if make_plots and pairs:
            n = len(pairs)
            fig, axes = plt.subplots(1, n, figsize=(5 * n, 4), squeeze=False)
            axes = axes[0]
        else:
            axes = [None] * len(pairs)

        for ax, (ra, rb) in zip(axes, pairs):
            va, vb = f"value_{ra}", f"value_{rb}"
            if va not in prot.columns or vb not in prot.columns:
                continue
            sub = prot[[va, vb]].dropna()
            if len(sub) < 3:
                continue
            diff = sub[va] - sub[vb]
            mean = (sub[va] + sub[vb]) / 2
            bias = float(diff.mean())
            sd = float(diff.std(ddof=1))
            loa_low, loa_high = bias - 1.96 * sd, bias + 1.96 * sd
            summary_rows.append({
                "Protocol": protocol,
                "Unit": unit,
                "Rater pair": f"{rater_name(ra)} vs {rater_name(rb)}",
                "n": len(sub),
                "Bias": round(bias, 4),
                "SD": round(sd, 4),
                "LoA_low": round(loa_low, 4),
                "LoA_high": round(loa_high, 4),
                "LoA_width": round(loa_high - loa_low, 4),
            })
            if ax is not None:
                ax.scatter(mean, diff, alpha=0.6, s=18)
                ax.axhline(bias, color="C0", lw=1.5)
                ax.axhline(loa_low, color="C3", ls="--", lw=1)
                ax.axhline(loa_high, color="C3", ls="--", lw=1)
                ax.set_title(f"{friendly_protocol(protocol)}\n{rater_name(ra)} vs {rater_name(rb)}")
                ax.set_xlabel(f"Mean ({unit})")
                ax.set_ylabel(f"Diff ({unit})")
        if make_plots and pairs:
            plt.tight_layout()
            plt.show()

    summary_df = pd.DataFrame(summary_rows)
    if len(summary_df):
        print("\n--- Bland–Altman bias summary ---")
        display(summary_df)
        print("\nPooled across pairs:")
        display(
            summary_df.groupby("Protocol")
            .agg(
                n_pairs=("Rater pair", "nunique"),
                mean_abs_bias=("Bias", lambda s: s.abs().mean()),
                mean_LoA_width=("LoA_width", "mean"),
            )
            .round(4)
            .reset_index()
        )
    return summary_df


# =============================================================================
# Pairwise redo tracker → analysis_outputs_v5/redo_tracker only
# =============================================================================

def _pair_pct_diff(val_a, val_b, protocol) -> float:
    va, vb = float(val_a), float(val_b)
    mid = 0.5 * (va + vb)
    floor = float(PROTOCOL_PCT_FLOOR.get(protocol, 1.0))
    return abs(va - vb) / max(abs(mid), floor)


def _style_tracker_sheet(ws, status_header="Status", highlight_header=None, highlight_value="Yes",
                         highlight_fill="FFF2CC"):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    pending_fill = PatternFill("solid", fgColor="FCE4D6")
    done_fill = PatternFill("solid", fgColor="C6EFCE")
    skip_fill = PatternFill("solid", fgColor="E7E6E6")
    focus_fill = PatternFill("solid", fgColor=highlight_fill)

    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 32
    status_col = headers.get(status_header)
    focus_col = headers.get(highlight_header) if highlight_header else None

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        if focus_col and ws.cell(r, focus_col).value == highlight_value:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = focus_fill
        if status_col:
            st = ws.cell(r, status_col)
            if st.value == "Pending":
                st.fill = pending_fill
            elif st.value == "Done":
                st.fill = done_fill
            elif st.value == "Skipped":
                st.fill = skip_fill

    for idx in range(1, ws.max_column + 1):
        letter = get_column_letter(idx)
        header = ws.cell(1, idx).value or ""
        width = 14
        if header in ("What to redo", "Why this was flagged", "Action"):
            width = 42
        elif header in ("Protocol", "Assigned to", "Compared with"):
            width = 22
        elif header in ("Notes",):
            width = 28
        elif header in ("Status", "Side", "Unit"):
            width = 10
        elif "Measurement" in str(header) or "Median" in str(header):
            width = 16
        ws.column_dimensions[letter].width = width


def _write_instructions_sheet(writer, title, bullets):
    from openpyxl.styles import Font, Alignment
    instr = pd.DataFrame({"How to use this tracker": [title, "", *bullets]})
    instr.to_excel(writer, sheet_name="Instructions", index=False)
    ws = writer.sheets["Instructions"]
    ws.column_dimensions["A"].width = 100
    ws["A1"].font = Font(bold=True, size=14)
    for r in range(1, ws.max_row + 1):
        ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 18


def build_pairwise_redo_tracker(
    comparison_df: pd.DataFrame,
    rater_keys: list[str],
    project_path: Path | str,
    exclude: Optional[list[str]] = None,
) -> pd.DataFrame:
    """
    Pairwise redo queue. Writes ONLY under analysis_outputs_v5/redo_tracker/.
    Never touches the live project-root redo_tracker.xlsx.
    """
    project_path = Path(project_path)
    live_mtime = None
    live = project_path / "redo_tracker.xlsx"
    if live.exists():
        live_mtime = live.stat().st_mtime

    keys = active_rater_keys(rater_keys, exclude)
    pairs = list(combinations(keys, 2))
    outlier_events = []

    for ra, rb in pairs:
        va, vb = f"value_{ra}", f"value_{rb}"
        if va not in comparison_df.columns or vb not in comparison_df.columns:
            continue
        pair_df = comparison_df.dropna(subset=[va, vb]).copy()
        if pair_df.empty:
            continue
        pair_df["diff"] = pair_df[va] - pair_df[vb]
        ra_name, rb_name = rater_name(ra), rater_name(rb)
        ra_email = ra.split(" | ")[1] if " | " in ra else ""
        rb_email = rb.split(" | ")[1] if " | " in rb else ""

        for protocol, grp in pair_df.groupby("true_protocol"):
            if len(grp) < REDO_MIN_PAIR_N:
                continue
            m_bias = grp["diff"].mean()
            std_bias = grp["diff"].std(ddof=1)
            if pd.notna(std_bias) and std_bias > 0:
                z_series = (grp["diff"] - m_bias) / std_bias
            else:
                z_series = pd.Series(np.nan, index=grp.index)

            unit = PROTOCOL_META.get(protocol, {}).get("unit", "")
            for idx_row, row in grp.iterrows():
                val_a, val_b = float(row[va]), float(row[vb])
                pct = _pair_pct_diff(val_a, val_b, protocol)
                z_score = float(z_series.loc[idx_row]) if idx_row in z_series.index else np.nan
                if pct < PAIR_PCT_MIN:
                    continue
                if pd.isna(z_score) or abs(z_score) < BA_Z_MIN:
                    continue

                case_mask = (
                    (comparison_df["patientId"] == row["patientId"])
                    & (comparison_df["laterality"] == row["laterality"])
                    & (comparison_df["true_protocol"] == protocol)
                )
                case_vals = {}
                for rk in keys:
                    col = f"value_{rk}"
                    if col in comparison_df.columns:
                        v = comparison_df.loc[case_mask, col]
                        if len(v) and pd.notna(v.iloc[0]):
                            case_vals[rk] = float(v.iloc[0])
                consensus = float(np.median(list(case_vals.values()))) if case_vals else np.nan

                candidates = [
                    (ra_name, ra_email, val_a, rb_name, val_b),
                    (rb_name, rb_email, val_b, ra_name, val_a),
                ]
                scored = []
                for name, email, their_val, other_name, other_val in candidates:
                    dist_c = abs(their_val - consensus) if pd.notna(consensus) else np.nan
                    other_d = abs(other_val - consensus) if pd.notna(consensus) else np.nan
                    scored.append((dist_c, other_d, name, email, their_val, other_name, other_val))

                dists = [s[0] for s in scored if pd.notna(s[0])]
                if dists:
                    max_dist = max(dists)
                    chosen = [s for s in scored if pd.notna(s[0]) and s[0] >= max_dist - 1e-12]
                else:
                    chosen = scored

                for dist_c, other_d, name, email, their_val, other_name, other_val in chosen:
                    suggested = (
                        "Yes" if pd.notna(other_d) and pd.notna(dist_c) and dist_c > other_d else
                        "Tie" if pd.notna(other_d) and pd.notna(dist_c) and dist_c == other_d else
                        "Yes"
                    )
                    outlier_events.append({
                        "assigned_rater": name,
                        "assigned_email": email,
                        "patientId": row["patientId"],
                        "laterality": row["laterality"],
                        "protocol": protocol,
                        "unit": unit,
                        "their_value": round(their_val, 4),
                        "other_rater": other_name,
                        "other_value": round(other_val, 4),
                        "difference_A_minus_B": round(float(row["diff"]), 4),
                        "pair_mean_bias": round(float(m_bias), 4) if pd.notna(m_bias) else np.nan,
                        "pair_std": round(float(std_bias), 4) if pd.notna(std_bias) else np.nan,
                        "z_score": round(z_score, 3) if pd.notna(z_score) else np.nan,
                        "pair_pct_diff": round(pct, 4),
                        "case_median_all_raters": round(consensus, 4) if pd.notna(consensus) else np.nan,
                        "abs_dist_from_median": round(dist_c, 4) if pd.notna(dist_c) else np.nan,
                        "suggested_focus": suggested,
                        "status": STATUS_DEFAULT,
                        "done_date": "",
                        "notes": "",
                    })

    redo_df = pd.DataFrame(outlier_events)
    if redo_df.empty:
        print(
            f"No redo tasks (≥{PAIR_PCT_MIN*100:.0f}% AND |z|≥{BA_Z_MIN}; "
            f"min pair n={REDO_MIN_PAIR_N})."
        )
        assert_live_tracker_untouched(project_path, live_mtime)
        return redo_df

    redo_df = redo_df.assign(_rank=redo_df["pair_pct_diff"].fillna(redo_df["z_score"].abs()))
    redo_df = (
        redo_df.sort_values(
            ["assigned_rater", "protocol", "patientId", "laterality", "_rank"],
            ascending=[True, True, True, True, False],
        )
        .drop_duplicates(subset=["assigned_rater", "patientId", "laterality", "protocol"], keep="first")
        .drop(columns="_rank")
        .reset_index(drop=True)
    )
    redo_df.insert(0, "redo_id", [f"R{i+1:04d}" for i in range(len(redo_df))])

    tracker_rows = []
    for _, r in redo_df.iterrows():
        focus = str(r.get("suggested_focus", "") or "")
        zabs = abs(float(r["z_score"])) if pd.notna(r.get("z_score")) else float("nan")
        pct = float(r["pair_pct_diff"]) if pd.notna(r.get("pair_pct_diff")) else float("nan")
        if focus == "Yes":
            action = "Priority: recheck YOUR landmarks — you are farther from the group median."
        elif focus == "Tie":
            action = "Recheck this case together with the other rater (both equally far from the median)."
        else:
            action = "Recheck this case (either rater may need a correction)."
        why = (
            f"You and {r['other_rater']} differ by ~{pct*100:.0f}% on {friendly_protocol(r['protocol'])} "
            f"(≥{PAIR_PCT_MIN*100:.0f}% threshold) and outside your usual Bland–Altman band "
            f"(|z|={zabs:.1f}). You are farther from the group median, so you recheck."
        )
        tracker_rows.append({
            "Task ID": r["redo_id"],
            "Status": r["status"],
            "Done date": r["done_date"],
            "Notes": r["notes"],
            "Assigned to": r["assigned_rater"],
            "Email": r["assigned_email"],
            "Patient ID": r["patientId"],
            "Side": friendly_side(r["laterality"]),
            "Protocol": friendly_protocol(r["protocol"]),
            "Unit": r["unit"],
            "What to redo": (
                f"Re-annotate {friendly_protocol(r['protocol'])} on patient {r['patientId']} "
                f"({friendly_side(r['laterality'])} side)."
            ),
            "Action": action,
            "Priority focus (you look like the outlier)?": (
                "Yes — start here" if focus == "Yes" else ("Tie" if focus == "Tie" else "No")
            ),
            "Your measurement": r["their_value"],
            "Compared with": r["other_rater"],
            "Their measurement": r["other_value"],
            "Group median (all raters)": r["case_median_all_raters"],
            "Why this was flagged": why,
            "Relative disagreement %": round(pct * 100, 1) if pd.notna(pct) else round(zabs, 2),
        })

    tracker_df = pd.DataFrame(tracker_rows)
    details_df = redo_df.rename(columns={
        "redo_id": "Task ID",
        "assigned_rater": "Assigned to",
        "assigned_email": "Email",
        "patientId": "Patient ID",
        "laterality": "Side",
        "protocol": "Protocol code",
        "unit": "Unit",
        "their_value": "Your measurement",
        "other_rater": "Compared with",
        "other_value": "Their measurement",
        "difference_A_minus_B": "Pair difference (A−B)",
        "pair_mean_bias": "Pair usual bias",
        "pair_std": "Pair SD",
        "z_score": "Z-score",
        "case_median_all_raters": "Group median",
        "abs_dist_from_median": "Distance from group median",
        "suggested_focus": "Suggested focus",
        "status": "Status",
        "done_date": "Done date",
        "notes": "Notes",
    })
    by_rater = (
        tracker_df.groupby(["Assigned to", "Email", "Status"], dropna=False)
        .size().reset_index(name="Number of tasks")
        .sort_values(["Assigned to", "Status"])
    )
    unique_cases = (
        tracker_df.groupby(["Patient ID", "Side", "Protocol"], as_index=False)
        .agg(**{
            "People assigned": ("Assigned to", "nunique"),
            "Assigned names": ("Assigned to", lambda s: ", ".join(sorted(s.unique()))),
            "Max relative disagreement %": ("Relative disagreement %", "max"),
        })
        .sort_values(["Protocol", "Patient ID", "Side"])
    )

    print("=== Redo tracker (pairwise) — v5 safe output folder ===")
    display(tracker_df.head(20))
    display(by_rater)

    out = redo_dir(project_path)
    csv_path = safe_write_path(project_path, REDO_SUBDIR_NAME, "redo_queue.csv")
    xlsx_path = safe_write_path(project_path, REDO_SUBDIR_NAME, "redo_tracker.xlsx")
    details_path = safe_write_path(project_path, REDO_SUBDIR_NAME, "redo_queue_details.csv")
    by_rater_path = safe_write_path(project_path, REDO_SUBDIR_NAME, "redo_queue_by_rater.csv")
    unique_path = safe_write_path(project_path, REDO_SUBDIR_NAME, "redo_queue_unique_cases.csv")

    tracker_df.to_csv(csv_path, index=False)
    details_df.to_csv(details_path, index=False)
    by_rater.to_csv(by_rater_path, index=False)
    unique_cases.to_csv(unique_path, index=False)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        _write_instructions_sheet(writer, "Redo tracker v5 (pairwise only)", [
            "This is a v5 DRAFT redo list for review — it lives in analysis_outputs_v5/redo_tracker/.",
            "It does NOT overwrite the live project-root redo_tracker.xlsx.",
            "",
            "Method: ≥10% relative gap AND outside pair Bland–Altman band (|z|≥1.96).",
            "Assigned to the rater farther from the case median (tie → both).",
            "",
            "HOW TO USE (after you approve promoting this draft):",
            "1. Open Pair disagreement → filter Assigned to your name.",
            "2. Re-annotate; set Status → Done.",
        ])
        tracker_df.to_excel(writer, sheet_name="Pair disagreement", index=False)
        by_rater.to_excel(writer, sheet_name="Pair — workload", index=False)
        unique_cases.to_excel(writer, sheet_name="Pair — unique cases", index=False)
        details_df.to_excel(writer, sheet_name="Pair — details", index=False)
        _style_tracker_sheet(
            writer.sheets["Pair disagreement"],
            status_header="Status",
            highlight_header="Priority focus (you look like the outlier)?",
            highlight_value="Yes — start here",
        )
        _style_tracker_sheet(writer.sheets["Pair — workload"], status_header="Status")

    assert_live_tracker_untouched(project_path, live_mtime)
    print(f"\nSaved DRAFT redo tracker → {xlsx_path}")
    print(f"Live tracker untouched → {live}")
    print(f"{len(tracker_df)} tasks · {tracker_df['Assigned to'].nunique()} people")
    return tracker_df


# =============================================================================
# JMP export (refined)
# =============================================================================

def export_jmp_long(
    comparison_df: pd.DataFrame,
    rater_keys: list[str],
    project_path: Path | str,
    exclude: Optional[list[str]] = None,
    include_single_rater: bool = False,
    all_rows: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """
    Primary analysis export for JMP.

    Refinements vs v4 notebook:
      • Uses true_protocol (keeps sulcus-angle and sulcus-angle-3cm separate)
      • Dedupes identical rows
      • Writes under analysis_outputs_v5/ (does not clobber live JMP at project root)
      • Optional: also export singles from raw results for coverage QA
    """
    keys = active_rater_keys(rater_keys, exclude)
    value_cols = [f"value_{rk}" for rk in keys if f"value_{rk}" in comparison_df.columns]

    src = comparison_df
    if include_single_rater and all_rows is not None:
        # Build a long table directly from protocol_result for full coverage QA
        res = all_rows[
            (all_rows["recordType"] == "protocol_result")
            & (all_rows["rater_key"].isin(keys))
        ].copy()
        res["laterality"] = res["laterality"].astype(str).str.strip().str.lower()
        res["patientId"] = res["patientId"].astype(str).str.strip()
        res["__group_timestamp"] = pd.to_numeric(
            res["groupId"].str.extract(r"-(\d+)$")[0], errors="coerce"
        ).fillna(0)
        res = (
            res.sort_values(["__group_timestamp", "__row_index"], ascending=False)
            .drop_duplicates(
                subset=["patientId", "laterality", "true_protocol", "rater_key"],
                keep="first",
            )
        )
        jmp_long = res.rename(columns={
            "rater_key": "Rater_ID",
            "resultValue": "Measurement_Value",
        })[["patientId", "laterality", "true_protocol", "Rater_ID", "Measurement_Value"]].copy()
        jmp_long["Rater_Name"] = jmp_long["Rater_ID"].map(rater_name)
        source_note = "all protocol_result rows (incl. single-rater)"
    else:
        jmp_long = comparison_df.melt(
            id_vars=["patientId", "laterality", "true_protocol"],
            value_vars=value_cols,
            var_name="Rater_ID",
            value_name="Measurement_Value",
        ).dropna(subset=["Measurement_Value"])
        jmp_long["Rater_ID"] = jmp_long["Rater_ID"].str.replace("^value_", "", regex=True)
        jmp_long["Rater_Name"] = jmp_long["Rater_ID"].map(rater_name)
        source_note = "matched ≥2-rater comparison_df"

    before = len(jmp_long)
    jmp_long = jmp_long.drop_duplicates(
        subset=["patientId", "laterality", "true_protocol", "Rater_ID"],
        keep="first",
    ).reset_index(drop=True)

    # Coverage check baked into export
    print(f"=== JMP long export ({source_note}) ===")
    print(f"Rows before dedupe: {before} → after: {len(jmp_long)}")
    print("Counts by true_protocol:")
    display(jmp_long.groupby("true_protocol").size().rename("n").to_frame())
    print("Counts by Rater_Name × true_protocol:")
    display(
        jmp_long.pivot_table(
            index="Rater_Name", columns="true_protocol",
            values="Measurement_Value", aggfunc="count", fill_value=0,
        ).astype(int)
    )

    out_path = safe_write_path(project_path, "jmp_statistical_comparison_long.csv")
    jmp_long.to_csv(out_path, index=False)
    print(f"Saved JMP export → {out_path}")
    print(f"(Live project-root JMP file was NOT overwritten.)")
    display(jmp_long.head())
    return jmp_long


# =============================================================================
# Planar agreement (optional, compact)
# =============================================================================

def planar_variability_summary(
    dataframes: dict[str, pd.DataFrame],
    rater_keys: list[str],
    project_path: Path | str | None = None,
    exclude: Optional[list[str]] = None,
) -> pd.DataFrame:
    from scipy.optimize import linear_sum_assignment

    keys = active_rater_keys(rater_keys, exclude)
    step_cols = [
        "patientId", "laterality", "true_protocol", "stepId", "stepLabel",
        "plane", "sliceIndex", "measurementType", "pointCount", "stepPointsMmJson",
    ]
    step_frames = {
        rk: dataframes[rk][dataframes[rk]["recordType"] == "step_measurement"][step_cols].copy()
        for rk in keys
    }
    merge_keys = ["patientId", "laterality", "true_protocol", "stepId"]

    def match_points_xy(pts_a, pts_b):
        n = len(pts_a)
        if n == 0 or n != len(pts_b):
            return pts_b
        a_xy = np.array([[p["x"], p["y"]] for p in pts_a])
        b_xy = np.array([[p["x"], p["y"]] for p in pts_b])
        cost = np.linalg.norm(a_xy[:, None, :] - b_xy[None, :, :], axis=2)
        _, col = linear_sum_assignment(cost)
        return [pts_b[i] for i in col]

    rows = []
    for ra, rb in combinations(keys, 2):
        merged = step_frames[ra].merge(step_frames[rb], on=merge_keys, suffixes=("_a", "_b"))
        for _, row in merged.iterrows():
            if row.get("pointCount_a") == 1:
                continue
            pts_a = parse_points_json(row["stepPointsMmJson_a"])[:2]
            pts_b = parse_points_json(row["stepPointsMmJson_b"])[:2]
            if not pts_a or len(pts_a) != len(pts_b):
                continue
            pts_b = match_points_xy(pts_a, pts_b)
            for p_a, p_b in zip(pts_a, pts_b):
                dist = float(np.linalg.norm([p_a["x"] - p_b["x"], p_a["y"] - p_b["y"]]))
                rows.append({
                    "true_protocol": row["true_protocol"],
                    "stepId": row["stepId"],
                    "rater_a": rater_name(ra),
                    "rater_b": rater_name(rb),
                    "dist_mm": dist,
                })

    planar_df = pd.DataFrame(rows)
    if planar_df.empty:
        print("No planar pairs found.")
        return planar_df

    stats = (
        planar_df.groupby(["true_protocol", "stepId"])["dist_mm"]
        .agg(["mean", "std", "max", "count"])
        .round(2)
        .reset_index()
    )
    print("=== Planar landmark variability (mm) ===")
    display(stats)

    if project_path is not None:
        path = safe_write_path(project_path, "planar_variability_summary.csv")
        stats.to_csv(path, index=False)
        print(f"Saved → {path}")
    return stats


# =============================================================================
# End-to-end pipeline
# =============================================================================

def run_pipeline(
    make_ba_plots: bool = True,
    run_planar: bool = True,
    jmp_include_single_rater: bool = True,
) -> dict:
    """Run the full v5 analysis. Safe: never writes live redo_tracker.xlsx."""
    ensure_packages()
    project_path = resolve_project_path()
    print(f"Project: {project_path}")
    print(f"v5 outputs → {project_path / OUTPUT_DIR_NAME}")

    live = project_path / "redo_tracker.xlsx"
    live_mtime = live.stat().st_mtime if live.exists() else None

    all_rows, dataframes, rater_keys = load_all_exports(project_path)
    dataframes = standardize_all(dataframes, rater_keys)
    # Rebuild all_rows from standardized frames for diagnostics consistency
    all_rows = pd.concat(dataframes.values(), ignore_index=True)

    diag = coverage_diagnostic(all_rows, active_rater_keys(rater_keys))
    # Persist diagnostic tables
    for name, df in diag.items():
        if isinstance(df, pd.DataFrame) and len(df):
            path = safe_write_path(project_path, "diagnostics", f"{name}.csv")
            df.to_csv(path)

    comparison_df, value_cols = build_comparison_df(dataframes, rater_keys)
    icc_summary, pearson_df = compute_icc_and_pearson(comparison_df, rater_keys)
    if len(icc_summary):
        icc_summary.to_csv(safe_write_path(project_path, "icc_summary.csv"), index=False)
    if len(pearson_df):
        pearson_df.to_csv(safe_write_path(project_path, "pearson_pairwise.csv"), index=False)

    bias_df = bland_altman_and_bias(comparison_df, rater_keys, make_plots=make_ba_plots)
    if len(bias_df):
        bias_df.to_csv(safe_write_path(project_path, "bland_altman_bias_summary.csv"), index=False)

    tracker_df = build_pairwise_redo_tracker(comparison_df, rater_keys, project_path)

    jmp_long = export_jmp_long(
        comparison_df, rater_keys, project_path,
        include_single_rater=jmp_include_single_rater,
        all_rows=all_rows,
    )

    planar_stats = pd.DataFrame()
    if run_planar:
        planar_stats = planar_variability_summary(dataframes, rater_keys, project_path)

    assert_live_tracker_untouched(project_path, live_mtime)
    print("\n✓ Pipeline complete. Live redo_tracker.xlsx untouched.")
    return {
        "project_path": project_path,
        "all_rows": all_rows,
        "dataframes": dataframes,
        "rater_keys": rater_keys,
        "comparison_df": comparison_df,
        "icc_summary": icc_summary,
        "pearson_df": pearson_df,
        "bias_df": bias_df,
        "tracker_df": tracker_df,
        "jmp_long": jmp_long,
        "planar_stats": planar_stats,
        "diagnostics": diag,
    }


if __name__ == "__main__":
    run_pipeline(make_ba_plots=False, run_planar=True)
