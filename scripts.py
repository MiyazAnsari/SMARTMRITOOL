# -*- coding: utf-8 -*-
"""Synced from scripts_v4.ipynb (notebook is the blueprint).
Keep scripts.py and scripts_v4.ipynb identical in analysis logic.
"""

"""
scripts_v4.py

Derived from scripts_v3.ipynb (Colab inter-rater reliability notebook).

# MSK Annotation Suite — Inter-Rater Reliability Notebook (Planar, Standardized)

**v4 additions:** automatic line endpoint standardization (direction-agnostic),
anatomical sulcus condyle relabeling by laterality, and a tidy landmark table
aligned by Case ID × Rater ID for ICC / ML export.

# MSK Annotation Suite — Inter-Rater Reliability Notebook (Planar)

Analyses inter-rater reliability for the **six knee-MRI measurement protocols** exported
by the MSK Annotation Suite (`protocolMeasurementCsv.ts`).

> **Scope: planar (in-plane) analysis only.** This notebook treats every landmark as a
> 2D coordinate (`x, y`) and deliberately ignores slice/Z-axis variability. A rater's
> chosen slice is taken as given; we ask only "given that two raters looked at (roughly)
> the same slice, how well do their in-plane clicks agree?" Z-axis disagreement,
> slice-selection consensus, and soft Z-targets for a 2-stage model are **out of scope**
> for this version and have been removed rather than left as dead code.

The notebook tells one continuous story, in order:

| Part | Section | Narrative beat |
|------|---------|----------------|
| I | 1 | **Ingest** — load every CSV, determine who the raters actually are |
| I | 1.5 | **Standardize** — normalize line directions; correct sulcus lateral/medial by laterality |
| I | 2 | **Inventory** — what did each rater annotate, and how much of it |
| II | 3 | **Clinical agreement** — do raters land on the same diagnostic number? (ICC, Pearson) |
| II | 4 | **Clinical agreement, visually** — Bland-Altman per protocol per rater pair |
| II | 5 | **Bias summary** — pooled bias / limits-of-agreement table, flagged for systematic offset |
| III | 6 | **Planar landmark agreement** — do raters click the same (x, y), independent of Z? |
| III | 7 | **Outlier triage** — which specific landmarks disagree the most, and why |
| IV | 8 | **Ground truth for ML** — average clinical values into one label, weighted by ICC |
| IV | 9 | **Readiness** — versioning hygiene, duplicate-run detection, per-protocol sample sizes |

Parts I–II answer "can we trust the *clinical numbers*"; Part III answers "can we trust
the *clicks themselves*, in-plane"; Part IV turns both answers into a usable dataset.

> **Rater identity:** a rater is identified by `(sessionUser, sessionUserEmail)`, **not**
> by filename. A single rater can export multiple CSVs (e.g. one per session); a single
> CSV file could in principle contain rows from more than one session. All files are
> loaded and concatenated first, then split into raters by identity.
>
> **Protocol identity:** `protocolId` stores `'sulcus-angle'` for **both** plain Sulcus
> Angle and Sulcus Angle 3 cm. The `groupId` prefix is the only reliable discriminator —
> all protocol grouping uses `true_protocol`, extracted from `groupId`.
>
> **Defaults to comparing all raters found.** Pairwise outputs (Bland-Altman, planar
> scatter) are produced for every rater pair; ICC and pooled summaries use all raters
> at once.

---
## Part I — Ingestion and Inventory

### 1. Mount Google Drive, Load CSVs, Identify Raters

Loads **every** CSV in the folder (regardless of filename), concatenates them, then
groups rows by `(sessionUser, sessionUserEmail)` to determine the actual rater set.
This means:
- One rater can have many files (e.g. split by session) — they're merged automatically.
- Filenames are irrelevant to rater identity; only used as a fallback display label if
  `sessionUser` is missing.
- The analysis defaults to **all raters found**, however many there are.
"""

import pandas as pd
import os
from pathlib import Path

try:
    from IPython.display import display  # type: ignore
except ImportError:
    def display(obj):
        print(obj)

# =============================================================================
# PATH CONFIG — identical in scripts_v4.ipynb and scripts.py
# =============================================================================
# Leave LOCAL_PROJECT_PATH as None to auto-detect Google Drive for Desktop.
# Or paste a Finder path, e.g.:
#   LOCAL_PROJECT_PATH = "/Users/you/Library/CloudStorage/GoogleDrive-you@email/My Drive/Current Knee MRI Project Folder"

# ── Ensure analysis packages exist (Colab + local) ──────────────────────────
import sys
import subprocess

def _ensure_packages(pkgs):
    missing = []
    for pkg, mod in pkgs:
        try:
            __import__(mod)
        except ImportError:
            missing.append(pkg)
    if missing:
        print('Installing:', ', '.join(missing))
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', '-q', *missing])

_ensure_packages([
    ('pingouin', 'pingouin'),
    ('openpyxl', 'openpyxl'),
    ('seaborn', 'seaborn'),
    ('scipy', 'scipy'),
    ('matplotlib', 'matplotlib'),
])

LOCAL_PROJECT_PATH = None
PROJECT_FOLDER_NAME = 'Current Knee MRI Project Folder'
CSV_SUBFOLDER_NAME = 'csv exports'


def _candidate_drive_project_paths():
    """Likely locations of the shared folder once Drive for Desktop is syncing."""
    home = Path.home()
    candidates = []
    if LOCAL_PROJECT_PATH:
        candidates.append(Path(LOCAL_PROJECT_PATH).expanduser())
    cloud = home / 'Library' / 'CloudStorage'
    if cloud.is_dir():
        for entry in sorted(cloud.iterdir()):
            if entry.name.startswith('GoogleDrive'):
                candidates.append(entry / 'My Drive' / PROJECT_FOLDER_NAME)
                candidates.append(entry / 'MyDrive' / PROJECT_FOLDER_NAME)
    for base in (
        home / 'Google Drive' / 'My Drive',
        home / 'Google Drive',
        Path('/Volumes/GoogleDrive/My Drive'),
    ):
        candidates.append(base / PROJECT_FOLDER_NAME)
    # Last resort: folder next to this file / notebook working dir
    try:
        here = Path(__file__).resolve().parent
    except NameError:
        here = Path.cwd()
    candidates.append(here / PROJECT_FOLDER_NAME)
    candidates.append(here)
    return candidates


def _in_colab():
    try:
        import google.colab  # type: ignore  # noqa: F401
        return True
    except ImportError:
        return False


def resolve_project_path():
    """Colab Drive mount if available; otherwise Drive Desktop / local folder."""
    if _in_colab():
        from google.colab import drive  # type: ignore
        drive.mount('/content/drive')
        project = Path('/content/drive/MyDrive') / PROJECT_FOLDER_NAME
        csv_dir = project / CSV_SUBFOLDER_NAME
        if not csv_dir.is_dir():
            raise FileNotFoundError(
                'Colab mounted Drive, but the project CSV folder was not found.\n\n'
                f'Expected: {csv_dir}\n\n'
                'Fix: In Google Drive (browser), open Shared with me → '
                f'"{PROJECT_FOLDER_NAME}" → Organize → Add shortcut to Drive → My Drive. '
                'Then Runtime → Restart session and re-run.'
            )
        return project

    for project in _candidate_drive_project_paths():
        if (project / CSV_SUBFOLDER_NAME).is_dir():
            return project
    searched = '\n'.join(f'  - {p / CSV_SUBFOLDER_NAME}' for p in _candidate_drive_project_paths())
    raise FileNotFoundError(
        'Could not find the measurement CSV folder.\n\n'
        f'Looking for: "{PROJECT_FOLDER_NAME}/{CSV_SUBFOLDER_NAME}"\n\n'
        'Fix checklist:\n'
        '  1. Install Google Drive for Desktop and sign in.\n'
        '  2. Add the shared project as a shortcut to My Drive.\n'
        '  3. Or set LOCAL_PROJECT_PATH near the top of this cell/file.\n\n'
        f'Paths tried:\n{searched}'
    )


project_path = str(resolve_project_path())
csv_folder = os.path.join(project_path, CSV_SUBFOLDER_NAME)
print(f'Using project folder: {project_path}')
print(f'Using CSV folder:     {csv_folder}')

if not os.path.exists(csv_folder):
    raise FileNotFoundError(f"CSV folder not found: {csv_folder}")

# ── Load every CSV file, tag with source filename and file mtime ───────────
csv_files = sorted(f for f in os.listdir(csv_folder) if f.lower().endswith('.csv'))
print(f"Found {len(csv_files)} CSV file(s)")

raw_frames = []
for fname in csv_files:
    fpath = os.path.join(csv_folder, fname)
    try:
        # Using engine='python' to more robustly handle 'on_bad_lines' and avoid C-buffer overflows
        fdf = pd.read_csv(fpath, on_bad_lines='warn', engine='python')
    except UnicodeDecodeError:
        try:
            # Fallback to cp1252 with error replacement for characters like 0x9d
            fdf = pd.read_csv(fpath, encoding='cp1252', encoding_errors='replace', on_bad_lines='warn', engine='python')
        except Exception as e:
            print(f"⚠ Skipping unreadable CSV '{fname}': {e}")
            continue
    except Exception as e:
        print(f"⚠ Skipping unreadable CSV '{fname}': {e}")
        continue

    # Skip non-export files (e.g. Pages docs renamed .csv)
    required_cols = {'groupId', 'recordType'}
    if not required_cols.issubset(set(fdf.columns)):
        print(
            f"⚠ Skipping '{fname}': missing required columns "
            f"{sorted(required_cols - set(fdf.columns))} (not a protocol measurement export)"
        )
        continue

    fdf['__source_file'] = fname
    # Capture file modification time to determine 'latest'
    fdf['__file_mtime'] = os.path.getmtime(fpath)
    # Capture original row index to handle ties within a single file
    fdf['__row_index'] = fdf.index
    raw_frames.append(fdf)

if not raw_frames:
    raise FileNotFoundError(f'No readable CSV files in {csv_folder}')

all_rows = pd.concat(raw_frames, ignore_index=True)
all_rows['true_protocol'] = all_rows['groupId'].str.extract(r'^([a-z][a-z0-9]*(?:-[a-z0-9]+)*?)-\d+$')

if 'sessionUserEmail' in all_rows.columns and all_rows['sessionUserEmail'].notna().any():
    id_cols = ['sessionUser', 'sessionUserEmail']
else:
    id_cols = ['__source_file']

all_rows['rater_key'] = all_rows[id_cols].astype(str).agg(' | '.join, axis=1)

dataframes = {key: grp.copy() for key, grp in all_rows.groupby('rater_key')}
rater_keys = sorted(dataframes.keys())
print(f"Raters identified: {len(rater_keys)}")

"""
### 1.5. Line Standardization — Endpoints and Sulcus Anatomical Matching

Raters may draw the same anatomical line in opposite directions (top↔bottom,
left↔right). Sulcus-angle raters may also swap lateral vs. medial condyle labels,
especially on left-knee axial slices. Before any agreement statistics, every
`step_measurement` row is normalized in-place:

1. **Endpoint standardization** — for each drawn line (`measurementType == 'distance'`,
   `pointCount ≥ 2`), sort pt[0]/pt[1] by geometry:
   - horizontal-ish (`|Δx| ≥ |Δy|`): left → right (ascending X)
   - vertical-ish (`|Δy| > |Δx|`): top → bottom (ascending Y)
2. **Sulcus condyle matching** — within each sulcus case, compare mean X of the
   two condyle lines and relabel by laterality:
   - **Left knee:** lateral condyle = higher X, medial = lower X
   - **Right knee:** lateral condyle = lower X, medial = higher X
3. **Tidy landmark export** — `standardized_landmarks_df` with one row per
   landmark endpoint, keyed on `patientId` (Case ID) × `rater_key` (Rater ID).
"""

import json

import numpy as np

# Protocols / steps subject to sulcus anatomical relabeling
SULCUS_PROTOCOLS = {'sulcus-angle', 'sulcus-angle-3cm'}
SULCUS_CONDYLE_STEPS = {'lateral-line', 'medial-line'}

# Sulcus grouping: one protocol run (both condyle lines share groupId)
SULCUS_GROUP_KEYS = ['patientId', 'laterality', 'true_protocol', 'groupId', 'rater_key']


def parse_points_json(val):
    """Parse stepPointsMmJson / pointsJson into a list of {x, y, z?} dicts."""
    try:
        pts = json.loads(val) if isinstance(val, str) else val
        return pts if isinstance(pts, list) else []
    except (TypeError, json.JSONDecodeError):
        return []


def serialize_points_json(pts):
    """Serialize point list back to the export JSON string format."""
    return json.dumps(pts)


def sync_flat_point_columns(row, pts):
    """Write standardized points back into p0_*/p1_*/p2_* flat columns."""
    for i, pt in enumerate(pts[:4]):
        for axis in ('x', 'y', 'z'):
            col = f'p{i}_{axis}'
            if col in row.index:
                row[col] = pt.get(axis, np.nan)
    return row


def line_center_x(pts):
    """Mean X of the two drawn endpoints (ignores foot-of-perpendicular)."""
    if len(pts) < 2:
        return np.nan
    return (pts[0]['x'] + pts[1]['x']) / 2.0


def standardize_line_endpoints(pts):
    """
    Direction-agnostic endpoint ordering for a drawn line segment.

    Rules (applied to pt[0] and pt[1] only):
      - Horizontal-ish (|Δx| >= |Δy|): sort left → right (ascending X).
      - Vertical-ish   (|Δy| >  |Δx|): sort top  → bottom (ascending Y).

    Any trailing derived points (e.g. foot-of-perpendicular at index 2) are
    preserved; when endpoints swap, the foot is re-anchored to the new pt[1].

    Returns
    -------
    (pts_out, swapped) : standardized list and whether pt[0]/pt[1] were reversed.
    """
    if len(pts) < 2:
        return pts, False

    p0, p1 = dict(pts[0]), dict(pts[1])
    dx = abs(p1['x'] - p0['x'])
    dy = abs(p1['y'] - p0['y'])

    swapped = False
    if dx >= dy:
        if p0['x'] > p1['x']:
            p0, p1 = p1, p0
            swapped = True
    else:
        if p0['y'] > p1['y']:
            p0, p1 = p1, p0
            swapped = True

    out = [p0, p1] + [dict(p) for p in pts[2:]]
    if swapped and len(out) >= 3:
        # Foot-of-perpendicular is tied to endpoint pt[1]; move it with the anchor.
        out[2]['x'] = out[1]['x']
        out[2]['y'] = out[1]['y']
        if 'z' in out[1]:
            out[2]['z'] = out[1]['z']
    return out, swapped


def lateral_is_higher_x(laterality):
    """
    Return True when the lateral condyle should sit on the higher-X image side.

    Left knee (axial, viewed from below): lateral on the right → higher X.
    Right knee: orientation flips → lateral on the left → lower X.
    """
    lat = str(laterality).strip().lower()
    if lat == 'left':
        return True
    if lat == 'right':
        return False
    return None


def apply_endpoint_standardization_to_row(row):
    """Standardize drawn-line endpoints on a single step_measurement row."""
    if row.get('recordType') != 'step_measurement':
        return row, False
    if row.get('measurementType') != 'distance':
        return row, False
    if pd.isna(row.get('pointCount')) or int(row['pointCount']) < 2:
        return row, False

    pts = parse_points_json(row.get('stepPointsMmJson'))
    if len(pts) < 2:
        return row, False

    std_pts, swapped = standardize_line_endpoints(pts)
    if not swapped:
        return row, False

    row = row.copy()
    row['stepPointsMmJson'] = serialize_points_json(std_pts)
    if 'pointsJson' in row.index and pd.notna(row.get('pointsJson')):
        row['pointsJson'] = serialize_points_json(std_pts)
    row = sync_flat_point_columns(row, std_pts)
    row['endpoints_standardized'] = True
    return row, True


def correct_sulcus_condyle_labels(df):
    """
    Within each sulcus protocol run, assign lateral-line / medial-line by mean X
    and laterality — regardless of the label the rater originally chose.

    Geometry stays on the same measurement row; only stepId / stepLabel are
    rewritten when a rater picked the wrong condyle name for that line's X position.
    """
    df = df.copy()
    if 'sulcus_label_corrected' not in df.columns:
        df['sulcus_label_corrected'] = False
    if 'original_stepId' not in df.columns:
        df['original_stepId'] = df['stepId']

    SULCUS_STEP_LABELS = {
        'lateral-line': 'Line from lateral condyle peak to sulcus',
        'medial-line':  'Line from medial condyle peak to sulcus',
    }

    step_mask = (
        (df['recordType'] == 'step_measurement')
        & (df['true_protocol'].isin(SULCUS_PROTOCOLS))
        & (df['stepId'].isin(SULCUS_CONDYLE_STEPS))
    )
    if not step_mask.any():
        return df

    for _group_key, grp in df.loc[step_mask].groupby(SULCUS_GROUP_KEYS):
        if len(grp) != 2:
            continue

        idx_a, idx_b = grp.index[0], grp.index[1]
        cx_a = line_center_x(parse_points_json(df.at[idx_a, 'stepPointsMmJson']))
        cx_b = line_center_x(parse_points_json(df.at[idx_b, 'stepPointsMmJson']))

        laterality = df.at[idx_a, 'laterality']
        higher_x_is_lateral = lateral_is_higher_x(laterality)
        if higher_x_is_lateral is None:
            continue

        # Map each row's geometry to the anatomically correct stepId
        if higher_x_is_lateral:
            target = {idx_a: ('lateral-line' if cx_a >= cx_b else 'medial-line'),
                      idx_b: ('lateral-line' if cx_b > cx_a else 'medial-line')}
        else:
            target = {idx_a: ('lateral-line' if cx_a <= cx_b else 'medial-line'),
                      idx_b: ('lateral-line' if cx_b < cx_a else 'medial-line')}

        for idx, new_step in target.items():
            if df.at[idx, 'stepId'] == new_step:
                continue
            df.at[idx, 'original_stepId'] = df.at[idx, 'stepId']
            df.at[idx, 'stepId'] = new_step
            if 'stepLabel' in df.columns:
                df.at[idx, 'stepLabel'] = SULCUS_STEP_LABELS.get(new_step, df.at[idx, 'stepLabel'])
            df.at[idx, 'sulcus_label_corrected'] = True

    return df


def standardize_rater_dataframe(df, rater_key):
    """
    Full standardization pass for one rater's annotations:
      1) endpoint ordering on all distance-type lines
      2) sulcus lateral/medial relabeling by laterality + mean X
    """
    df = df.copy()
    df['rater_key'] = rater_key
    df['endpoints_standardized'] = False
    df['original_stepId'] = df['stepId']

    step_idx = df.index[df['recordType'] == 'step_measurement']
    n_endpoints = 0
    for idx in step_idx:
        new_row, changed = apply_endpoint_standardization_to_row(df.loc[idx])
        if changed:
            df.loc[idx] = new_row
            n_endpoints += 1

    df = correct_sulcus_condyle_labels(df)
    n_sulcus = int(df.get('sulcus_label_corrected', pd.Series(False)).sum())

    print(
        f"  {rater_key.split(' | ')[0]}: "
        f"{n_endpoints} line(s) endpoint-standardized, "
        f"{n_sulcus} sulcus condyle row(s) anatomically relabeled"
    )
    return df


def build_standardized_landmarks_df(dataframes, rater_keys):
    """
    Tidy landmark table for inter-rater variability / ML.

    One row per drawn endpoint (or single point), aligned by:
      - patientId  (Case ID)
      - rater_key  (Rater ID)
      - laterality, true_protocol, stepId (standardized)

    Line landmarks expose start_x/start_y and end_x/end_y; multi-point steps
    also include point_index for finer-grained ICC.
    """
    rows = []
    for rk in rater_keys:
        df = dataframes[rk]
        steps = df[df['recordType'] == 'step_measurement']
        for _, row in steps.iterrows():
            pts = parse_points_json(row.get('stepPointsMmJson'))
            if not pts:
                continue

            mtype = row.get('measurementType')
            pc = int(row['pointCount']) if pd.notna(row.get('pointCount')) else len(pts)
            # Planar comparison uses pt[0:2] for 3-point distance lines
            effective = pts[:2] if (mtype == 'distance' and pc == 3) else pts

            base = {
                'patientId':     row['patientId'],
                'laterality':    row['laterality'],
                'true_protocol': row['true_protocol'],
                'groupId':       row.get('groupId'),
                'rater_key':     rk,
                'rater_name':    rk.split(' | ')[0],
                'stepId':        row['stepId'],
                'original_stepId': row.get('original_stepId', row['stepId']),
                'stepLabel':     row.get('stepLabel'),
                'plane':         row.get('plane'),
                'sliceIndex':    row.get('sliceIndex'),
                'measurementType': mtype,
                'pointCount':    pc,
                'endpoints_standardized': bool(row.get('endpoints_standardized', False)),
                'sulcus_label_corrected': bool(row.get('sulcus_label_corrected', False)),
            }

            if len(effective) == 2 and mtype == 'distance':
                rows.append({
                    **base,
                    'landmark_role': 'line',
                    'point_index': 0,
                    'start_x': effective[0]['x'],
                    'start_y': effective[0]['y'],
                    'start_z': effective[0].get('z', np.nan),
                    'end_x':   effective[1]['x'],
                    'end_y':   effective[1]['y'],
                    'end_z':   effective[1].get('z', np.nan),
                    'x': effective[0]['x'],
                    'y': effective[0]['y'],
                    'z': effective[0].get('z', np.nan),
                })
            else:
                for i, pt in enumerate(effective):
                    rows.append({
                        **base,
                        'landmark_role': 'point',
                        'point_index': i,
                        'start_x': np.nan, 'start_y': np.nan, 'start_z': np.nan,
                        'end_x': np.nan,   'end_y': np.nan,   'end_z': np.nan,
                        'x': pt['x'],
                        'y': pt['y'],
                        'z': pt.get('z', np.nan),
                    })

    landmarks = pd.DataFrame(rows)
    if landmarks.empty:
        return landmarks

    # Stable sort for inspection / pivoting
    sort_cols = ['patientId', 'laterality', 'true_protocol', 'stepId', 'point_index', 'rater_key']
    sort_cols = [c for c in sort_cols if c in landmarks.columns]
    return landmarks.sort_values(sort_cols).reset_index(drop=True)


# ── Apply standardization to every rater, rebuild dataframes ─────────────────
print("\nApplying line standardization (endpoints + sulcus anatomical matching)...")
dataframes = {
    rk: standardize_rater_dataframe(dataframes[rk], rk)
    for rk in rater_keys
}

standardized_landmarks_df = build_standardized_landmarks_df(dataframes, rater_keys)
landmarks_out = os.path.join(project_path, 'standardized_landmarks.csv')
standardized_landmarks_df.to_csv(landmarks_out, index=False)
print(f"\nStandardized landmark table: {len(standardized_landmarks_df)} rows")
print(f"  Cases: {standardized_landmarks_df['patientId'].nunique()}")
print(f"  Raters: {standardized_landmarks_df['rater_key'].nunique()}")
print(f"  Saved to: {landmarks_out}")
display(standardized_landmarks_df.head(10))

"""
### 2. Preview Annotations
Record types, true protocol coverage, and column layout per rater — a sanity check
before any statistics are computed.
"""

for rk in rater_keys:
    df = dataframes[rk]
    print(f"\n{'='*60}")
    print(f"Rater: {rk}  |  Shape: {df.shape}")
    print(f"recordType     : {df['recordType'].value_counts().to_dict()}")
    print(f"true_protocol  : {df['true_protocol'].value_counts().to_dict()}")
    display(df[df['recordType'] == 'protocol_result'].head(3))

"""
---
## Part II — Clinical Value Agreement

Before trusting any single landmark click, we first ask the coarser question: do raters
agree on the **final clinical number** each protocol produces (an angle, a ratio, a
distance in mm)? This is the standard inter-rater reliability question and is answered
with ICC and Bland-Altman, exactly as it would be for any manual measurement — Z-axis
variability does not enter into it.

### 3. Build the Comparison Table, Then Measure Agreement

Builds one wide table (`comparison_df`) keyed on `patientId + laterality + true_protocol`,
with one column per rater's `resultValue`. If a rater re-ran the same case, their duplicate
rows are collapsed to the **median** value before merge so each annotator contributes at
most one measurement per case.

Reports, in order of preference:
1. **Per-protocol ICC(2,1)** across all raters jointly — the primary clinical metric.
2. **Per-protocol pairwise Pearson r** — secondary, pairwise-complete-case view.
3. **Pooled Pearson matrix** — reference only; mixes ratios, degrees, and mm across
   protocols, so it is not clinically interpretable on its own.
"""

import sys
import subprocess
from itertools import combinations
import numpy as np

try:
    import pingouin as pg
except ImportError:
    subprocess.run([sys.executable, '-m', 'pip', 'install', 'pingouin', '-q'])
    import pingouin as pg

# User defined reliable composite key for deduplication
CASE_KEY = [
    'recordType', 'patientId', 'sessionUser', 'sessionUserEmail',
    'laterality', 'sequenceName', 'plane', 'protocolId'
]

def extract_results(df, rater_key):
    """Extract protocol_result rows using a robust composite key, selecting the latest measurement via groupId timestamp."""
    result_df = df[df['recordType'] == 'protocol_result'].copy()

    # Extract Unix timestamp from the end of the groupId (e.g., protocol-id-1712345678000)
    # We use (\\d+) digits at the end of the string
    result_df['__group_timestamp'] = pd.to_numeric(
        result_df['groupId'].str.extract(r'-(\d+)$')[0],
        errors='coerce'
    ).fillna(0)

    # Sort by the embedded timestamp (desc) then original row index (desc) to get latest first
    sub = result_df.sort_values(['__group_timestamp', '__row_index'], ascending=False)

    # Drop duplicates using the comprehensive composite key, keeping the first (latest)
    before_len = len(sub)
    sub = sub.drop_duplicates(subset=CASE_KEY, keep='first')
    n_dup = before_len - len(sub)

    if n_dup > 0:
        print(f"  {rater_key.split(' | ')[0]}: ignored {n_dup} older duplicate measurement(s) based on groupId timestamp")

    return sub[['patientId', 'laterality', 'true_protocol', 'resultValue', 'resultUnit']].rename(
        columns={'resultValue': f'value_{rater_key}', 'resultUnit': f'unit_{rater_key}'}
    )

print("Extracting protocol_result rows (selecting latest measurement using groupId timestamps)... ")
processed = [extract_results(dataframes[k], k) for k in rater_keys]

# Re-bind JOIN_KEYS for the outer merge operation
JOIN_KEYS = ['patientId', 'laterality', 'true_protocol']

comparison_df = processed[0]
for nxt in processed[1:]:
    comparison_df = comparison_df.merge(nxt, on=JOIN_KEYS, how='outer')

value_cols = [f'value_{k}' for k in rater_keys]
comparison_df['n_raters_present'] = comparison_df[value_cols].notna().sum(axis=1)
comparison_df = comparison_df[comparison_df['n_raters_present'] >= 2].reset_index(drop=True)

print(f"\nMatched records (≥2 raters): {len(comparison_df)}")
display(comparison_df.groupby('true_protocol')[value_cols].count())

# Used by Bland-Altman, bias summary, redo queue, and planar cells
rater_pairs = list(combinations(rater_keys, 2))
print(f"Rater pairs: {len(rater_pairs)}")

"""
### 3.1. Statistical Reliability Metrics (ICC and Pearson r)

We calculate **ICC(2,1)** (two-way random effects, absolute agreement, single rater) to assess the reliability of the clinical values across all raters. We also compute pairwise **Pearson correlation coefficients** for each rater pair per protocol.
"""

icc_results = []
pearson_results = []

# Get list of protocols present in the comparison dataframe
protocols = sorted(comparison_df['true_protocol'].dropna().unique())

for protocol in protocols:
    # Subset data for this protocol
    prot_df = comparison_df[comparison_df['true_protocol'] == protocol].copy()
    if prot_df.empty:
        continue

    # 1. ICC Calculation
    # Convert wide to long format for pingouin
    long_df = prot_df.melt(
        id_vars=['patientId', 'laterality'],
        value_vars=value_cols,
        var_name='rater',
        value_name='value'
    ).dropna()

    # Create a unique subject ID for the ICC
    long_df['subject'] = long_df['patientId'].astype(str) + '_' + long_df['laterality'].astype(str)

    # Filter to subjects that have at least 2 ratings for this specific protocol
    subject_counts = long_df['subject'].value_counts()
    valid_subjects = subject_counts[subject_counts >= 2].index
    long_df_balanced = long_df[long_df['subject'].isin(valid_subjects)]

    if long_df_balanced['rater'].nunique() > 1 and len(valid_subjects) >= 5:
        try:
            # Using nan_policy='omit' to handle unbalanced data
            try:
                icc = pg.intraclass_corr(
                    data=long_df_balanced, targets='subject', raters='rater',
                    ratings='value', nan_policy='omit',
                )
            except TypeError:
                # Older pingouin (common on Colab) has no nan_policy=
                icc = pg.intraclass_corr(
                    data=long_df_balanced, targets='subject', raters='rater',
                    ratings='value',
                )
            icc['Protocol'] = protocol
            icc_results.append(icc)
        except Exception as e:
            print(f"Could not calculate ICC for {protocol}: {e}")

    # 2. Pairwise Pearson r
    for ra, rb in combinations(rater_keys, 2):
        v_a, v_b = f'value_{ra}', f'value_{rb}'
        pair_sub = prot_df[[v_a, v_b]].dropna()
        if len(pair_sub) > 2:
            r = pair_sub[v_a].corr(pair_sub[v_b])
            pearson_results.append({
                'Protocol': protocol,
                'Rater A': ra.split(' | ')[0],
                'Rater B': rb.split(' | ')[0],
                'n': len(pair_sub),
                'Pearson r': round(r, 4)
            })

# Display ICC Results
if icc_results:
    icc_df = pd.concat(icc_results)
    # Select Absolute Agreement (ICC2) regardless of label convention
    icc_summary = icc_df[icc_df['Type'].str.contains('ICC2|ICC\\(A,1\\)')].copy()
    print("\n--- Inter-Rater Reliability: ICC(2,1) per Protocol ---")
    ci_col = 'CI95' if 'CI95' in icc_summary.columns else 'CI95%'
    cols_to_show = ['Protocol', 'Type', 'ICC', 'F', 'df1', 'df2', 'pval', ci_col]
    display(icc_summary[cols_to_show].round(3))

# Display Pearson Results
if pearson_results:
    pearson_df = pd.DataFrame(pearson_results)
    print("\n--- Pairwise Pearson Correlation (Mean per Protocol) ---")
    display(pearson_df.groupby('Protocol')['Pearson r'].mean().to_frame())

"""
### 6. Planar Coordinate Variability and ICC

We now measure how consistent the raw **(x, y) landmark clicks** are in the imaging plane, using the Hungarian algorithm to match points and calculating pairwise planar distances.
"""

import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import pandas as pd
from scipy.optimize import linear_sum_assignment
from itertools import combinations

# Define missing dependencies needed by this cell
step_cols = ['patientId', 'laterality', 'true_protocol', 'stepId', 'stepLabel',
             'plane', 'sliceIndex', 'measurementType', 'pointCount',
             'stepPointsMmJson', 'pointsJson']
step_frames = {
    rk: dataframes[rk][dataframes[rk]['recordType'] == 'step_measurement'][step_cols].copy()
    for rk in rater_keys
}
STEP_MERGE_KEYS = ['patientId', 'laterality', 'true_protocol', 'stepId']

def match_points_xy(pts_a, pts_b):
    n = len(pts_a)
    if n == 0 or n != len(pts_b):
        return pts_b
    a_xy = np.array([[p['x'], p['y']] for p in pts_a])
    b_xy = np.array([[p['x'], p['y']] for p in pts_b])
    cost_matrix = np.linalg.norm(a_xy[:, np.newaxis, :] - b_xy[np.newaxis, :, :], axis=2)
    row_ind, col_ind = linear_sum_assignment(cost_matrix)
    return [pts_b[i] for i in col_ind]

def planar_distance_for_pair(ra, rb):
    merged = step_frames[ra].merge(step_frames[rb], on=STEP_MERGE_KEYS, suffixes=('_a', '_b'))
    rows = []
    for _, row in merged.iterrows():
        pts_a = parse_points_json(row['stepPointsMmJson_a'])[:2]
        pts_b = parse_points_json(row['stepPointsMmJson_b'])[:2]
        if len(pts_a) != len(pts_b) or len(pts_a) == 0: continue
        pts_b_matched = match_points_xy(pts_a, pts_b)
        for i, (p_a, p_b) in enumerate(zip(pts_a, pts_b_matched)):
            dist = np.linalg.norm([p_a['x'] - p_b['x'], p_a['y'] - p_b['y']])
            rows.append({'rater_a': ra, 'rater_b': rb, 'stepId': row['stepId'], 'dist_mm': dist})
    return pd.DataFrame(rows)

rater_pairs = list(combinations(rater_keys, 2))
pairwise_planar_dfs = [planar_distance_for_pair(ra, rb) for ra, rb in rater_pairs]
planar_variability_df = pd.concat(pairwise_planar_dfs, ignore_index=True)

display(planar_variability_df.groupby('stepId')['dist_mm'].agg(['mean', 'std', 'max']).round(2))

planar_stats = planar_variability_df.groupby('stepId')['dist_mm'].agg(['mean', 'std', 'max', 'count']).round(2)
planar_out_path = os.path.join(project_path, 'planar_variability_summary.csv')
planar_stats.to_csv(planar_out_path)
print(f"Planar variability summary saved to: {planar_out_path}")
display(planar_stats)

"""
### 4. Clinical Agreement, Visually — Bland-Altman per Protocol

**Bland-Altman is inherently pairwise** (it plots difference vs. mean for two raters),
so with N raters it is produced for every pair: $\\binom{N}{2}$ figures, each with one
row per protocol. With many raters this can be a lot of plots — restrict
`RATER_PAIRS_OVERRIDE` in Section 3 if needed.
"""

import matplotlib.pyplot as plt
import seaborn as sns

sns.set_theme(style='whitegrid')

# Reference ranges from MSK README per true_protocol
PROTOCOL_META = {
    'tt-tg':             {'unit': 'mm',    'normal': '<15 mm',   'borderline': '15–20 mm'},
    'insall-salvati':    {'unit': 'ratio', 'normal': '0.8–1.2',  'borderline': None},
    'patellar-tilt':     {'unit': '°',     'normal': '<10°',     'borderline': '10–20°'},
    'sulcus-angle':      {'unit': '°',     'normal': '<145°',    'borderline': None},
    'sulcus-angle-3cm':  {'unit': '°',     'normal': '<145°',    'borderline': None},
    'caton-deschamps':   {'unit': 'ratio', 'normal': '0.6–1.3',  'borderline': None},
}

protocols = sorted(comparison_df['true_protocol'].dropna().unique())

for ra, rb in rater_pairs:
    # Use the deduplicated comparison_df
    pair_df = comparison_df.dropna(subset=[f'value_{ra}', f'value_{rb}']).copy()
    if pair_df.empty:
        continue

    pair_df['diff'] = pair_df[f'value_{ra}'] - pair_df[f'value_{rb}']

    n_prot    = len(protocols)
    fig, axes = plt.subplots(n_prot, 2, figsize=(14, 5 * n_prot), squeeze=False)
    fig.suptitle(f'Rater comparison (Deduplicated): {ra} vs {rb}', fontsize=14, y=1.0)

    for ax_row, protocol in zip(axes, protocols):
        sub = pair_df[pair_df['true_protocol'] == protocol]
        if sub.empty:
            ax_row[0].axis('off'); ax_row[1].axis('off')
            continue

        diff = sub['diff']
        mean = sub[[f'value_{ra}', f'value_{rb}']].mean(axis=1)
        md_, sd_ = diff.mean(), diff.std()
        meta = PROTOCOL_META.get(protocol, {})
        unit = meta.get('unit', '')

        sns.histplot(diff, kde=True, ax=ax_row[0], color='steelblue')
        ax_row[0].axvline(0, color='red', linestyle='--', linewidth=1.5)
        ax_row[0].set_title(f'{protocol} [{unit}]\nDifference Distribution')
        ax_row[0].set_xlabel(f'{ra} − {rb}')

        ax_row[1].scatter(mean, diff, alpha=0.6, color='steelblue')
        for level, label, ls, color in [
            (md_,              f'Bias: {md_:.3f}',               "--", "red"),
            (md_ + 1.96 * sd_, f"+1.96 SD: {md_+1.96*sd_:.3f}", ":",  "gray"),
            (md_ - 1.96 * sd_, f"−1.96 SD: {md_-1.96*sd_:.3f}", ":",  "gray"),
        ]:
            ax_row[1].axhline(level, linestyle=ls, color=color, label=label)

        subtitle = f"Normal: {meta['normal']}" if meta.get('normal') else ''
        ax_row[1].set_title(f'{protocol}\nBland-Altman (n={len(sub)})\n{subtitle}')
        ax_row[1].set_xlabel('Mean of pair')
        ax_row[1].set_ylabel(f'{ra} − {rb} [{unit}]')
        ax_row[1].legend(fontsize=8)

    plt.tight_layout()
    plt.show()

"""
### 5. Bias Summary — Pooled Bland-Altman Statistics

The numeric counterpart to Section 4: bias and 95% limits-of-agreement (LoA) per
`true_protocol`, computed per rater pair and then pooled. Ratios (CDI, IS) have
`resultUnit = NaN` — handled correctly via `PROTOCOL_META`. Pairs with a large bias
relative to LoA width are flagged as a likely systematic rater offset rather than
random noise.
"""

summary_rows = []
for ra, rb in rater_pairs:
    pair_df = comparison_df.dropna(subset=[f'value_{ra}', f'value_{rb}']).copy()
    if pair_df.empty:
        continue
    pair_df['diff'] = pair_df[f'value_{ra}'] - pair_df[f'value_{rb}']
    for protocol, grp in pair_df.groupby('true_protocol'):
        d    = grp['diff']
        meta = PROTOCOL_META.get(protocol, {})
        summary_rows.append({
            'Protocol':   protocol,
            'Rater pair': f"{ra} vs {rb}",
            'Unit':       meta.get('unit', ''),
            'n':          len(grp),
            'Bias':       round(d.mean(), 4),
            'SD':         round(d.std(),  4),
            'LoA_lower':  round(d.mean() - 1.96 * d.std(), 4),
            'LoA_upper':  round(d.mean() + 1.96 * d.std(), 4),
            'LoA_width':  round(1.96 * 2 * d.std(), 4),
        })

summary_df = pd.DataFrame(summary_rows)
print("Bland-Altman Statistics by True Protocol and Rater Pair:")
display(summary_df)

for _, row in summary_df.iterrows():
    if abs(row['Bias']) > abs(row['LoA_width']) * 0.3:
        print(f"⚠  {row['Protocol']} ({row['Rater pair']}): large bias ({row['Bias']:.3f}) "
              f"relative to LoA width — check for systematic rater offset.")

# ── Pooled (protocol-level only, all pairs combined) ──────────────────────────
print("\nPooled across all rater pairs:")
display(summary_df.groupby('Protocol').agg(
    n_pairs=('Rater pair', 'nunique'),
    mean_abs_bias=('Bias', lambda s: s.abs().mean()),
    mean_LoA_width=('LoA_width', 'mean'),
).round(4).reset_index())

"""
### 5.1 Redo Queue — Pair Disagreement (≥10% + Bland–Altman)

Builds a **per-person redo list** when two raters disagree in a way that is both clinically meaningful and unusual for that pair:

- Require `|A − B| / scale ≥ 10%` (protocol floor on the scale so near-zero values don’t explode).
- **And** require the pair difference outside that pair’s Bland–Altman band (`|z| ≥ 1.96` vs the pair’s mean bias/SD).
- **Only the rater farther from the case median** is assigned (ties → both).

This dual gate avoids the old “|z| alone” over-flagging (tiny BA SD) and also avoids listing every ≥10% gap when the pair’s usual scatter already covers it.

**Output:** pair pages inside combined `redo_tracker.xlsx` (plus `redo_queue.csv`). Run §5.2 afterward to merge consensus pages into the same workbook.

"""

import sys
import subprocess
try:
    import openpyxl  # noqa: F401
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', '-q', 'openpyxl'])

import re
from pathlib import Path


# ── Human-readable tracker helpers ──────────────────────────────────────────
PROTOCOL_DISPLAY = {
    'tt-tg': 'TT-TG distance',
    'insall-salvati': 'Insall–Salvati ratio',
    'patellar-tilt': 'Patellar tilt',
    'sulcus-angle': 'Sulcus angle',
    'sulcus-angle-3cm': 'Sulcus angle (3 cm)',
    'caton-deschamps': 'Caton–Deschamps ratio',
}

def _friendly_protocol(p):
    return PROTOCOL_DISPLAY.get(str(p), str(p))

def _friendly_side(s):
    s = str(s).strip().lower()
    return {'left': 'Left', 'right': 'Right'}.get(s, str(s).title())

def _style_tracker_sheet(ws, status_header='Status', highlight_header=None, highlight_value='Yes',
                         highlight_fill='FFF2CC'):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    pending_fill = PatternFill('solid', fgColor='FCE4D6')
    done_fill = PatternFill('solid', fgColor='C6EFCE')
    skip_fill = PatternFill('solid', fgColor='E7E6E6')
    focus_fill = PatternFill('solid', fgColor=highlight_fill)

    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[1].height = 32

    status_col = headers.get(status_header)
    focus_col = headers.get(highlight_header) if highlight_header else None

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical='center')
        if focus_col and ws.cell(r, focus_col).value == highlight_value:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = focus_fill
        if status_col:
            st = ws.cell(r, status_col)
            if st.value == 'Pending':
                st.fill = pending_fill
            elif st.value == 'Done':
                st.fill = done_fill
            elif st.value == 'Skipped':
                st.fill = skip_fill

    # Reasonable column widths
    for idx in range(1, ws.max_column + 1):
        letter = get_column_letter(idx)
        header = ws.cell(1, idx).value or ''
        width = 14
        if header in ('What to redo', 'Why this was flagged', 'Action'):
            width = 42
        elif header in ('Protocol', 'Assigned to', 'Compared with', 'Partner who agreed'):
            width = 22
        elif header in ('Notes',):
            width = 28
        elif header in ('Status', 'Side', 'Unit'):
            width = 10
        elif 'Measurement' in str(header) or 'Median' in str(header):
            width = 16
        ws.column_dimensions[letter].width = width

def _write_instructions_sheet(writer, title, bullets):
    instr = pd.DataFrame({
        'How to use this tracker': [
            title,
            '',
            *bullets,
        ]
    })
    instr.to_excel(writer, sheet_name='Instructions', index=False)
    ws = writer.sheets['Instructions']
    ws.column_dimensions['A'].width = 100
    from openpyxl.styles import Font, Alignment
    ws['A1'].font = Font(bold=True, size=14)
    for r in range(1, ws.max_row + 1):
        ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[r].height = 18

# ── Config ──────────────────────────────────────────────────────────────────
# Pairwise redo: require a CLINICALLY meaningful gap (≈10% relative), not only a
# high z-score. Bland–Altman bias is often small; tiny SD makes |z|>2/3 flag many
# cases that are still clinically close. Aligns with mentor 10% rule.
PAIR_PCT_MIN = 0.10               # |A−B| / scale ≥ 10%
BA_Z_MIN = 1.96                   # also outside this pair's Bland–Altman band
REDO_MIN_PAIR_N = 5               # skip unstable pair/protocol groups
STATUS_DEFAULT = 'Pending'        # Pending | Done | Skipped

# Denominator floors so near-zero values (e.g. tilt≈0) don't inflate % gaps
PROTOCOL_PCT_FLOOR_51 = {
    'tt-tg': 10.0,             # mm
    'patellar-tilt': 5.0,      # degrees
    'sulcus-angle': 100.0,     # degrees
    'sulcus-angle-3cm': 100.0,
    'insall-salvati': 0.8,     # ratio
    'caton-deschamps': 0.8,
}

def _pair_pct_diff(val_a, val_b, protocol):
    """Relative disagreement between two raters on one case."""
    va, vb = float(val_a), float(val_b)
    mid = 0.5 * (va + vb)
    floor = float(PROTOCOL_PCT_FLOOR_51.get(protocol, 1.0))
    denom = max(abs(mid), floor)
    return abs(va - vb) / denom

# ── Build outlier events (pair × protocol × case) ───────────────────────────
outlier_events = []

for ra, rb in rater_pairs:
    va, vb = f'value_{ra}', f'value_{rb}'
    pair_df = comparison_df.dropna(subset=[va, vb]).copy()
    if pair_df.empty:
        continue

    pair_df['diff'] = pair_df[va] - pair_df[vb]
    ra_name, rb_name = ra.split(' | ')[0], rb.split(' | ')[0]
    ra_email = ra.split(' | ')[1] if ' | ' in ra else ''
    rb_email = rb.split(' | ')[1] if ' | ' in rb else ''

    for protocol, grp in pair_df.groupby('true_protocol'):
        if len(grp) < REDO_MIN_PAIR_N:
            continue
        m_bias = grp['diff'].mean()
        std_bias = grp['diff'].std(ddof=1)
        # z used with BA_Z_MIN dual gate below (plus 10% relative)
        if pd.notna(std_bias) and std_bias > 0:
            z_series = (grp['diff'] - m_bias) / std_bias
        else:
            z_series = pd.Series(np.nan, index=grp.index)

        meta = PROTOCOL_META.get(protocol, {})
        unit = meta.get('unit', '')

        for idx_row, row in grp.iterrows():
            val_a = float(row[va])
            val_b = float(row[vb])
            pct = _pair_pct_diff(val_a, val_b, protocol)
            z_score = float(z_series.loc[idx_row]) if idx_row in z_series.index else np.nan
            # Dual gate: clinically meaningful AND unusual for this pair (BA band)
            if pct < PAIR_PCT_MIN:
                continue
            if pd.isna(z_score) or abs(z_score) < BA_Z_MIN:
                continue

            # Case MEDIAN across all raters present
            case_mask = (
                (comparison_df['patientId'] == row['patientId'])
                & (comparison_df['laterality'] == row['laterality'])
                & (comparison_df['true_protocol'] == protocol)
            )
            case_vals = {}
            for rk in rater_keys:
                col = f'value_{rk}'
                if col in comparison_df.columns:
                    v = comparison_df.loc[case_mask, col]
                    if len(v) and pd.notna(v.iloc[0]):
                        case_vals[rk] = float(v.iloc[0])
            consensus = float(np.median(list(case_vals.values()))) if case_vals else np.nan

            # Only the rater farther from the group MEDIAN redoes (tie → both)
            candidates = [
                (ra_name, ra_email, val_a, rb_name, val_b),
                (rb_name, rb_email, val_b, ra_name, val_a),
            ]
            scored = []
            for rater_name, rater_email, their_val, other_name, other_val in candidates:
                dist_consensus = abs(their_val - consensus) if pd.notna(consensus) else np.nan
                other_dist = abs(other_val - consensus) if pd.notna(consensus) else np.nan
                scored.append((dist_consensus, other_dist, rater_name, rater_email, their_val, other_name, other_val))

            dists = [s[0] for s in scored if pd.notna(s[0])]
            if dists:
                max_dist = max(dists)
                chosen = [s for s in scored if pd.notna(s[0]) and s[0] >= max_dist - 1e-12]
            else:
                # no group median — assign both when 10% pair gap
                chosen = scored

            for dist_consensus, other_dist, rater_name, rater_email, their_val, other_name, other_val in chosen:
                suggested = (
                    'Yes' if pd.notna(other_dist) and pd.notna(dist_consensus) and dist_consensus > other_dist else
                    'Tie' if pd.notna(other_dist) and pd.notna(dist_consensus) and dist_consensus == other_dist else
                    'Yes'
                )
                outlier_events.append({
                    'assigned_rater': rater_name,
                    'assigned_email': rater_email,
                    'patientId': row['patientId'],
                    'laterality': row['laterality'],
                    'protocol': protocol,
                    'unit': unit,
                    'their_value': round(their_val, 4),
                    'other_rater': other_name,
                    'other_value': round(other_val, 4),
                    'difference_A_minus_B': round(float(row['diff']), 4),
                    'pair_mean_bias': round(float(m_bias), 4) if pd.notna(m_bias) else np.nan,
                    'pair_std': round(float(std_bias), 4) if pd.notna(std_bias) else np.nan,
                    'z_score': round(z_score, 3) if pd.notna(z_score) else np.nan,
                    'pair_pct_diff': round(pct, 4),
                    'case_median_all_raters': round(consensus, 4) if pd.notna(consensus) else np.nan,
                    'abs_dist_from_median': round(dist_consensus, 4) if pd.notna(dist_consensus) else np.nan,
                    'suggested_focus': suggested,
                    'status': STATUS_DEFAULT,
                    'done_date': '',
                    'notes': '',
                })

redo_df = pd.DataFrame(outlier_events)


if redo_df.empty:
    print(
        f'No clinical redo tasks found '
        f'(≥{PAIR_PCT_MIN*100:.0f}% relative gap AND |z|≥{BA_Z_MIN} vs pair BA; '
        f'furthest from case MEDIAN; min n={REDO_MIN_PAIR_N}).'
    )
else:
    # One task per person × case × protocol (same case may appear in multiple pairs)
    redo_df = redo_df.assign(_rank=redo_df['pair_pct_diff'].fillna(redo_df['z_score'].abs()))
    redo_df = (
        redo_df.sort_values(
            ['assigned_rater', 'protocol', 'patientId', 'laterality', '_rank'],
            ascending=[True, True, True, True, False],
        )
        .drop_duplicates(
            subset=['assigned_rater', 'patientId', 'laterality', 'protocol'],
            keep='first',
        )
        .drop(columns='_rank')
        .reset_index(drop=True)
    )
    redo_df.insert(0, 'redo_id', [f'R{i+1:04d}' for i in range(len(redo_df))])

    # ── Build rater-friendly tracker view ───────────────────────────────────
    tracker_rows = []
    for _, r in redo_df.iterrows():
        focus = str(r.get('suggested_focus', '') or '')
        zabs = abs(float(r['z_score'])) if pd.notna(r.get('z_score')) else float('nan')
        pct = float(r['pair_pct_diff']) if pd.notna(r.get('pair_pct_diff')) else float('nan')
        if focus == 'Yes':
            action = 'Priority: recheck YOUR landmarks — you are farther from the group median.'
        elif focus == 'Tie':
            action = 'Recheck this case together with the other rater (both equally far from the median).'
        else:
            action = 'Recheck this case (either rater may need a correction).'
        why = (
            f"You and {r['other_rater']} differ by ~{pct*100:.0f}% on {_friendly_protocol(r['protocol'])} "
            f"(≥{PAIR_PCT_MIN*100:.0f}% threshold) and outside your usual Bland–Altman band "
            f"(|z|={zabs:.1f}). You are farther from the group median, so you recheck."
        )
        tracker_rows.append({
            'Task ID': r['redo_id'],
            'Status': r['status'],
            'Done date': r['done_date'],
            'Notes': r['notes'],
            'Assigned to': r['assigned_rater'],
            'Email': r['assigned_email'],
            'Patient ID': r['patientId'],
            'Side': _friendly_side(r['laterality']),
            'Protocol': _friendly_protocol(r['protocol']),
            'Unit': r['unit'],
            'What to redo': (
                f"Re-annotate { _friendly_protocol(r['protocol']) } on patient {r['patientId']} "
                f"({_friendly_side(r['laterality'])} side)."
            ),
            'Action': action,
            'Priority focus (you look like the outlier)?': (
                'Yes — start here' if focus == 'Yes' else ('Tie' if focus == 'Tie' else 'No')
            ),
            'Your measurement': r['their_value'],
            'Compared with': r['other_rater'],
            'Their measurement': r['other_value'],
            'Group median (all raters)': r['case_median_all_raters'],
            'Why this was flagged': why,
            'Relative disagreement %': round(float(r['pair_pct_diff'])*100, 1) if 'pair_pct_diff' in r.index and pd.notna(r.get('pair_pct_diff')) else round(zabs, 2),
        })

    tracker_df = pd.DataFrame(tracker_rows)

    # Keep technical raw table on a Details sheet
    details_df = redo_df.rename(columns={
        'redo_id': 'Task ID',
        'assigned_rater': 'Assigned to',
        'assigned_email': 'Email',
        'patientId': 'Patient ID',
        'laterality': 'Side',
        'protocol': 'Protocol code',
        'unit': 'Unit',
        'their_value': 'Your measurement',
        'other_rater': 'Compared with',
        'other_value': 'Their measurement',
        'difference_A_minus_B': 'Pair difference (A−B)',
        'pair_mean_bias': 'Pair usual bias',
        'pair_std': 'Pair SD',
        'z_score': 'Z-score',
        'case_median_all_raters': 'Group median',
        'abs_dist_from_median': 'Distance from group median',
        'suggested_focus': 'Suggested focus',
        'status': 'Status',
        'done_date': 'Done date',
        'notes': 'Notes',
    })

    by_rater = (
        tracker_df.groupby(['Assigned to', 'Email', 'Status'], dropna=False)
        .size()
        .reset_index(name='Number of tasks')
        .sort_values(['Assigned to', 'Status'])
    )
    unique_cases = (
        tracker_df.groupby(['Patient ID', 'Side', 'Protocol'], as_index=False)
        .agg(
            **{
                'People assigned': ('Assigned to', 'nunique'),
                'Assigned names': ('Assigned to', lambda s: ', '.join(sorted(s.unique()))),
                'Max relative disagreement %': ('Relative disagreement %', 'max'),
            }
        )
        .sort_values(['Protocol', 'Patient ID', 'Side'])
    )

    print('=== Redo tracker (pair-disagreement queue) ===')
    display(tracker_df.head(20))
    print('\n=== Workload by person ===')
    display(by_rater)
    print('\n=== Unique cases ===')
    display(unique_cases)
    print(
        f"\n{len(tracker_df)} tasks · {tracker_df['Assigned to'].nunique()} people · "
        f"{len(unique_cases)} unique case/protocol rows"
    )

    out_dir = Path(project_path)
    csv_path = out_dir / 'redo_queue.csv'
    xlsx_path = out_dir / 'redo_tracker.xlsx'

    # Persist pair tracker for §5.2 combined workbook rebuild
    tracker_df.to_csv(csv_path, index=False)
    details_df.to_csv(out_dir / 'redo_queue_details.csv', index=False)
    by_rater.to_csv(out_dir / 'redo_queue_by_rater.csv', index=False)
    unique_cases.to_csv(out_dir / 'redo_queue_unique_cases.csv', index=False)

    try:
        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            _write_instructions_sheet(writer, 'Combined redo tracker', [
                'This workbook has TWO task lists (different methods):',
                '  • "Pair disagreement" (§5.1) — unusual disagreement between two specific raters.',
                '  • "Consensus redo" (§5.2) — outside an agreeing cluster (plus Difficult/Calibration group pages).',
                '',
                'How to work:',
                '1. Open "Pair disagreement" and/or "Consensus redo".',
                '2. Filter "Assigned to" to your name.',
                '3. Follow What to redo + Action. Mark Status = Done when finished.',
                '4. "Workload by person" summarizes both lists.',
                '5. Ignore "Details …" sheets unless you are analyzing the stats.',
                '',
                'Note: Re-run §5.2 after §5.1 to refresh the full combined workbook.',
                '§5.1 alone writes the pair pages; §5.2 merges both into this file.',
            ])
            tracker_df.to_excel(writer, sheet_name='Pair disagreement', index=False)
            by_rater.to_excel(writer, sheet_name='Pair — workload', index=False)
            unique_cases.to_excel(writer, sheet_name='Pair — unique cases', index=False)
            details_df.to_excel(writer, sheet_name='Pair — details', index=False)
            _style_tracker_sheet(
                writer.sheets['Pair disagreement'],
                status_header='Status',
                highlight_header='Priority focus (you look like the outlier)?',
                highlight_value='Yes — start here',
            )
            _style_tracker_sheet(writer.sheets['Pair — workload'], status_header='Status')
        print(f'\nSaved combined tracker (pair pages): {xlsx_path}')
        print('(Run §5.2 next to add Consensus pages into the same file.)')
    except ImportError:
        print('\n⚠ openpyxl not installed — wrote CSV only.')

    print(f'Saved: {csv_path}')
    print(
        '\nOpen redo_tracker.xlsx → Instructions, then "Pair disagreement". '
        'After §5.2 runs, Consensus pages appear in the same workbook.'
    )

"""
### 5.2 Consensus Redo Queue — Majority Cluster (+ Difficult / Calibration)

Complements §5.1. Replaces furthest-from-median (which always picked someone when
values were merely spread out).

**Redo (cluster outsider):**
1. On each `patientId + laterality + protocol` with ≥3 raters, find the largest
   **agreeing cluster** (every member within `AGREE`% of each other on the protocol scale).
2. Cluster must meet min size: **2 if n=3**, else **max(3, ⌈n/2⌉−1)** (scales to 5–6 raters).
3. Cluster median = median of cluster members only. Assign redo to raters **outside**
   the cluster who are **≥10%** from that median.

**No valid cluster (or two opposing camps):**
- **Calibration / style conflict** — case ordering matches known protocol-level rater
  high/low tendencies (systematic camps, not a one-off hard case).
- **Difficult** — residual ambiguity / tricky patient; **no redo assigned**.

**Output:** same `redo_tracker.xlsx` with Consensus redo + Difficult + Calibration sheets.
"""

import sys
import subprocess
try:
    import openpyxl  # noqa: F401
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', '-q', 'openpyxl'])

from pathlib import Path
from itertools import combinations
import math


# ── Human-readable tracker helpers ──────────────────────────────────────────
PROTOCOL_DISPLAY = {
    'tt-tg': 'TT-TG distance',
    'insall-salvati': 'Insall–Salvati ratio',
    'patellar-tilt': 'Patellar tilt',
    'sulcus-angle': 'Sulcus angle',
    'sulcus-angle-3cm': 'Sulcus angle (3 cm)',
    'caton-deschamps': 'Caton–Deschamps ratio',
}

def _friendly_protocol(p):
    return PROTOCOL_DISPLAY.get(str(p), str(p))

def _friendly_side(s):
    s = str(s).strip().lower()
    return {'left': 'Left', 'right': 'Right'}.get(s, str(s).title())

def _style_tracker_sheet(ws, status_header='Status', highlight_header=None, highlight_value='Yes',
                         highlight_fill='FFF2CC'):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    pending_fill = PatternFill('solid', fgColor='FCE4D6')
    done_fill = PatternFill('solid', fgColor='C6EFCE')
    skip_fill = PatternFill('solid', fgColor='E7E6E6')
    focus_fill = PatternFill('solid', fgColor=highlight_fill)

    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[1].height = 32

    status_col = headers.get(status_header)
    focus_col = headers.get(highlight_header) if highlight_header else None

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical='center')
        if focus_col and ws.cell(r, focus_col).value == highlight_value:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = focus_fill
        if status_col:
            st = ws.cell(r, status_col)
            if st.value == 'Pending':
                st.fill = pending_fill
            elif st.value == 'Done':
                st.fill = done_fill
            elif st.value == 'Skipped':
                st.fill = skip_fill

    for idx in range(1, ws.max_column + 1):
        letter = get_column_letter(idx)
        header = ws.cell(1, idx).value or ''
        width = 14
        if header in ('What to redo', 'Why this was flagged', 'Action', 'Notes', 'Why marked'):
            width = 42
        elif header in ('Protocol', 'Assigned to', 'Compared with', 'Cluster members', 'High-tendency raters', 'Low-tendency raters'):
            width = 22
        elif header in ('Status', 'Side', 'Unit'):
            width = 10
        elif 'Measurement' in str(header) or 'Median' in str(header):
            width = 16
        ws.column_dimensions[letter].width = width

def _write_instructions_sheet(writer, title, bullets):
    instr = pd.DataFrame({'How to use this tracker': [title, '', *bullets]})
    instr.to_excel(writer, sheet_name='Instructions', index=False)
    ws = writer.sheets['Instructions']
    ws.column_dimensions['A'].width = 100
    from openpyxl.styles import Font, Alignment
    ws['A1'].font = Font(bold=True, size=14)
    for r in range(1, ws.max_row + 1):
        ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[r].height = 18

# ── Config ──────────────────────────────────────────────────────────────────
CONSENSUS_MIN_RATERS = 3
CLUSTER_AGREE_PCT = 0.05          # members of a cluster agree within 5%
CLUSTER_OUT_PCT = 0.10            # outsider redo if ≥10% from cluster median
STYLE_BIAS_MIN = 0.05             # |mean signed % vs case median| to call a rater high/low
STYLE_MATCH_FRAC = 0.60           # fraction of raters matching their usual side → Calibration
STATUS_DEFAULT_52 = 'Pending'

PROTOCOL_PCT_FLOOR = {
    'tt-tg': 10.0,
    'patellar-tilt': 5.0,
    'sulcus-angle': 100.0,
    'sulcus-angle-3cm': 100.0,
    'insall-salvati': 0.8,
    'caton-deschamps': 0.8,
}

def _protocol_scale(values, protocol):
    floor = float(PROTOCOL_PCT_FLOOR.get(protocol, 1.0))
    return max(abs(float(np.median(values))), floor)

def _pct_diff(value, ref, protocol=None):
    ref = float(ref)
    value = float(value)
    floor = float(PROTOCOL_PCT_FLOOR.get(protocol, 1.0)) if protocol is not None else 1.0
    denom = max(abs(ref), floor)
    return abs(value - ref) / denom

def _min_cluster_size(n):
    """Scale cluster requirement with annotator count (3 now → 5–6 later)."""
    if n <= 3:
        return 2
    return max(3, int(math.ceil(n / 2) - 1))  # 4→3, 5→3, 6→3

def _largest_agreeing_clusters(vals_by_name, protocol):
    """Return non-overlapping clusters (largest first) with internal spread ≤ AGREE%."""
    names = list(vals_by_name)
    n = len(names)
    sc = _protocol_scale(list(vals_by_name.values()), protocol)
    found = []
    for k in range(n, 1, -1):
        for comb in combinations(names, k):
            vs = [vals_by_name[x] for x in comb]
            spread = (max(vs) - min(vs)) / sc
            if spread <= CLUSTER_AGREE_PCT:
                found.append((set(comb), spread))
        if found:
            break
    if not found:
        return [], sc
    # Prefer larger, then tighter; greedily take non-overlapping
    found.sort(key=lambda t: (-len(t[0]), t[1]))
    chosen = []
    used = set()
    for cluster, spread in found:
        if cluster.isdisjoint(used):
            chosen.append((cluster, spread))
            used |= cluster
    return chosen, sc

# ── Build per-case wide values ──────────────────────────────────────────────
case_keys_52 = ['patientId', 'laterality', 'true_protocol']
wide = comparison_df[case_keys_52 + value_cols].copy()

# Pass 1: protocol-level rater tendency (signed % vs case median) for Calibration routing
_tendency_rows = []
for _, row in wide.iterrows():
    protocol = row['true_protocol']
    present = []
    for rk in rater_keys:
        col = f'value_{rk}'
        if col in row.index and pd.notna(row[col]):
            present.append((rk.split(' | ')[0], float(row[col])))
    if len(present) < CONSENSUS_MIN_RATERS:
        continue
    vals = [v for _, v in present]
    med = float(np.median(vals))
    sc = _protocol_scale(vals, protocol)
    for name, val in present:
        _tendency_rows.append({
            'protocol': protocol,
            'rater': name,
            'signed_pct': (val - med) / sc,
        })

tendency_df = pd.DataFrame(_tendency_rows)
if len(tendency_df):
    rater_bias = (
        tendency_df.groupby(['protocol', 'rater'])['signed_pct']
        .agg(mean_signed_pct='mean', n_cases='count')
        .reset_index()
    )
else:
    rater_bias = pd.DataFrame(columns=['protocol', 'rater', 'mean_signed_pct', 'n_cases'])

bias_lookup = {
    (r['protocol'], r['rater']): float(r['mean_signed_pct'])
    for _, r in rater_bias.iterrows()
}

# Pass 2: classify each case
redo_events = []
difficult_events = []
calibration_events = []
case_outcomes = []

for _, row in wide.iterrows():
    protocol = row['true_protocol']
    present = []  # (rk, name, email, val)
    for rk in rater_keys:
        col = f'value_{rk}'
        if col in row.index and pd.notna(row[col]):
            name = rk.split(' | ')[0]
            email = rk.split(' | ')[1] if ' | ' in rk else ''
            present.append((rk, name, email, float(row[col])))
    n = len(present)
    if n < CONSENSUS_MIN_RATERS:
        continue

    vals_by_name = {name: val for _, name, _, val in present}
    email_by_name = {name: email for _, name, email, _ in present}
    unit = PROTOCOL_META.get(protocol, {}).get('unit', '')
    need = _min_cluster_size(n)
    clusters, sc = _largest_agreeing_clusters(vals_by_name, protocol)
    valid = [(c, sp) for c, sp in clusters if len(c) >= need]
    all_vals = list(vals_by_name.values())
    case_median = float(np.median(all_vals))
    case_spread_pct = (max(all_vals) - min(all_vals)) / sc

    base_case = {
        'patientId': row['patientId'],
        'laterality': row['laterality'],
        'protocol': protocol,
        'unit': unit,
        'n_raters_on_case': n,
        'case_median_all': round(case_median, 4),
        'case_spread_pct': round(100.0 * case_spread_pct, 1),
        'min_cluster_needed': need,
        'best_cluster_size': len(clusters[0][0]) if clusters else 1,
    }

    # Two opposing valid camps → Difficult (split), not forced redo
    if len(valid) >= 2:
        meds = [float(np.median([vals_by_name[x] for x in c])) for c, _ in valid[:2]]
        if abs(meds[0] - meds[1]) / sc >= CLUSTER_OUT_PCT:
            camp_a = ', '.join(sorted(valid[0][0]))
            camp_b = ', '.join(sorted(valid[1][0]))
            difficult_events.append({
                **base_case,
                'mark_type': 'split_camps',
                'cluster_median': np.nan,
                'cluster_members': f'{camp_a} || {camp_b}',
                'why': (
                    f'Two agreeing camps differ by ≥{CLUSTER_OUT_PCT*100:.0f}% '
                    f'({camp_a} vs {camp_b}). Escalate / discuss — do not assign personal redo.'
                ),
            })
            case_outcomes.append({**base_case, 'outcome': 'Difficult_split'})
            continue

    if len(valid) >= 1:
        cluster, spread = valid[0]
        cmed = float(np.median([vals_by_name[x] for x in cluster]))
        outsiders = [name for name in vals_by_name if name not in cluster]
        any_redo = False
        for name in outsiders:
            val = vals_by_name[name]
            pct = _pct_diff(val, cmed, protocol)
            if pct < CLUSTER_OUT_PCT:
                continue
            any_redo = True
            redo_events.append({
                **base_case,
                'assigned_rater': name,
                'assigned_email': email_by_name.get(name, ''),
                'their_value': round(val, 4),
                'cluster_median': round(cmed, 4),
                'cluster_members': ', '.join(sorted(cluster)),
                'cluster_spread_pct': round(100.0 * spread, 1),
                'pct_from_cluster': round(100.0 * pct, 1),
                'reason': (
                    f'Outside agreeing cluster [{", ".join(sorted(cluster))}] '
                    f'(cluster median={cmed:.4f}, you={val:.4f}, '
                    f'{pct*100:.1f}% ≥ {CLUSTER_OUT_PCT*100:.0f}% threshold)'
                ),
                'status': STATUS_DEFAULT_52,
                'done_date': '',
                'notes': '',
            })
        case_outcomes.append({
            **base_case,
            'outcome': 'Redo' if any_redo else 'Consensus_ok',
            'cluster_members': ', '.join(sorted(cluster)),
        })
        continue

    # No valid cluster — Calibration vs Difficult
    signs_match = []
    high_names, low_names = [], []
    for name, val in vals_by_name.items():
        bias = bias_lookup.get((protocol, name), 0.0)
        case_sign = 0 if abs(val - case_median) / sc < 1e-9 else (1 if val > case_median else -1)
        bias_sign = 0 if abs(bias) < STYLE_BIAS_MIN else (1 if bias > 0 else -1)
        if bias_sign != 0:
            if bias_sign > 0:
                high_names.append(name)
            else:
                low_names.append(name)
            if case_sign != 0:
                signs_match.append(case_sign == bias_sign)

    match_frac = (sum(signs_match) / len(signs_match)) if signs_match else 0.0
    strong_style = (
        len(signs_match) >= 2
        and match_frac >= STYLE_MATCH_FRAC
        and (len(high_names) >= 1 and len(low_names) >= 1)
    )

    if strong_style:
        calibration_events.append({
            **base_case,
            'mark_type': 'style_conflict',
            'high_tendency_raters': ', '.join(sorted(set(high_names))),
            'low_tendency_raters': ', '.join(sorted(set(low_names))),
            'style_match_pct': round(100.0 * match_frac, 0),
            'why': (
                f'No agreeing cluster (need ≥{need} within {CLUSTER_AGREE_PCT*100:.0f}%). '
                f'Case order matches known protocol style camps '
                f'(high: {", ".join(sorted(set(high_names))) or "—"}; '
                f'low: {", ".join(sorted(set(low_names))) or "—"}). '
                f'Landmark calibration / training — not a one-off difficult patient.'
            ),
        })
        case_outcomes.append({**base_case, 'outcome': 'Calibration'})
    else:
        difficult_events.append({
            **base_case,
            'mark_type': 'no_cluster',
            'cluster_median': np.nan,
            'cluster_members': '',
            'why': (
                f'No agreeing cluster of ≥{need} raters within {CLUSTER_AGREE_PCT*100:.0f}% '
                f'(spread {100*case_spread_pct:.0f}%). Likely tricky / ambiguous landmarks — '
                f'no personal redo assigned.'
            ),
        })
        case_outcomes.append({**base_case, 'outcome': 'Difficult'})

redo_df = pd.DataFrame(redo_events)
difficult_df = pd.DataFrame(difficult_events)
calibration_df = pd.DataFrame(calibration_events)
outcomes_df = pd.DataFrame(case_outcomes)

print('=== §5.2 case outcomes (cluster consensus) ===')
if len(outcomes_df):
    print(outcomes_df['outcome'].value_counts().to_string())
    print('\nBy protocol:')
    print(pd.crosstab(outcomes_df['protocol'], outcomes_df['outcome']).to_string())
else:
    print('No multi-rater cases.')

# Protocol-level calibration summary (who tends high/low)
calib_summary_rows = []
for protocol, gbias in rater_bias.groupby('protocol'):
    highs = gbias[gbias['mean_signed_pct'] >= STYLE_BIAS_MIN].sort_values('mean_signed_pct', ascending=False)
    lows = gbias[gbias['mean_signed_pct'] <= -STYLE_BIAS_MIN].sort_values('mean_signed_pct')
    n_cal = int((calibration_df['protocol'] == protocol).sum()) if len(calibration_df) else 0
    n_diff = int((difficult_df['protocol'] == protocol).sum()) if len(difficult_df) else 0
    n_redo = int((redo_df['protocol'] == protocol).sum()) if len(redo_df) else 0
    calib_summary_rows.append({
        'Protocol': _friendly_protocol(protocol),
        'High-tendency raters': ', '.join(
            f"{r['rater']} ({100*r['mean_signed_pct']:+.0f}%)" for _, r in highs.iterrows()
        ) or '—',
        'Low-tendency raters': ', '.join(
            f"{r['rater']} ({100*r['mean_signed_pct']:+.0f}%)" for _, r in lows.iterrows()
        ) or '—',
        'Calibration cases': n_cal,
        'Difficult cases': n_diff,
        'Consensus redo tasks': n_redo,
        'Note': (
            'Style camps present — prefer landmark review over mass redo'
            if (len(highs) and len(lows) and n_cal >= 3)
            else ('Mostly tight agreement' if n_diff + n_cal == 0 else 'Mixed')
        ),
    })
calib_summary_df = pd.DataFrame(calib_summary_rows)

# ── Build tracker views ─────────────────────────────────────────────────────
out_dir = Path(project_path)

if len(redo_df):
    redo_df = redo_df.sort_values(
        ['assigned_rater', 'protocol', 'patientId', 'laterality', 'pct_from_cluster'],
        ascending=[True, True, True, True, False],
    ).reset_index(drop=True)
    redo_df.insert(0, 'redo_id', [f'C{i+1:04d}' for i in range(len(redo_df))])

    tracker_rows = []
    for _, r in redo_df.iterrows():
        tracker_rows.append({
            'Task ID': r['redo_id'],
            'Status': r['status'],
            'Done date': r['done_date'],
            'Notes': r['notes'],
            'Assigned to': r['assigned_rater'],
            'Email': r['assigned_email'],
            'Patient ID': r['patientId'],
            'Side': _friendly_side(r['laterality']),
            'Protocol': _friendly_protocol(r['protocol']),
            'Unit': r['unit'],
            'What to redo': (
                f"Re-annotate {_friendly_protocol(r['protocol'])} on patient {r['patientId']} "
                f"({_friendly_side(r['laterality'])} side)."
            ),
            'Issue type': 'Outside agreeing cluster',
            'Action': (
                'Recheck YOUR landmarks — a clear group agreed with each other and you are outside that cluster.'
            ),
            'Your measurement': r['their_value'],
            'Cluster median': r['cluster_median'],
            'Cluster members': r['cluster_members'],
            '# raters on this case': r['n_raters_on_case'],
            '% away from cluster': r['pct_from_cluster'],
            'Why this was flagged': r['reason'],
        })
    tracker_df = pd.DataFrame(tracker_rows)
else:
    tracker_df = pd.DataFrame()

if len(difficult_df):
    difficult_df = difficult_df.sort_values(
        ['protocol', 'patientId', 'laterality']
    ).reset_index(drop=True)
    difficult_df.insert(0, 'case_id', [f'D{i+1:04d}' for i in range(len(difficult_df))])
    difficult_tracker = pd.DataFrame([{
        'Case ID': r['case_id'],
        'Status': 'Open',
        'Notes': '',
        'Patient ID': r['patientId'],
        'Side': _friendly_side(r['laterality']),
        'Protocol': _friendly_protocol(r['protocol']),
        'Unit': r['unit'],
        'Mark type': 'Split camps' if r['mark_type'] == 'split_camps' else 'No agreeing cluster',
        '# raters': r['n_raters_on_case'],
        'Case spread %': r['case_spread_pct'],
        'Best cluster size': r['best_cluster_size'],
        'Min cluster needed': r['min_cluster_needed'],
        'Camps / members': r.get('cluster_members', ''),
        'Action': 'Do NOT assign personal redo — review as a group / mentor if needed.',
        'Why marked': r['why'],
    } for _, r in difficult_df.iterrows()])
else:
    difficult_tracker = pd.DataFrame()

if len(calibration_df):
    calibration_df = calibration_df.sort_values(
        ['protocol', 'patientId', 'laterality']
    ).reset_index(drop=True)
    calibration_df.insert(0, 'case_id', [f'K{i+1:04d}' for i in range(len(calibration_df))])
    calibration_tracker = pd.DataFrame([{
        'Case ID': r['case_id'],
        'Status': 'Open',
        'Notes': '',
        'Patient ID': r['patientId'],
        'Side': _friendly_side(r['laterality']),
        'Protocol': _friendly_protocol(r['protocol']),
        'Unit': r['unit'],
        'Mark type': 'Style / calibration conflict',
        '# raters': r['n_raters_on_case'],
        'Case spread %': r['case_spread_pct'],
        'High-tendency raters': r['high_tendency_raters'],
        'Low-tendency raters': r['low_tendency_raters'],
        'Style match %': r['style_match_pct'],
        'Action': (
            'Landmark calibration — discuss consistent high/low camps for this protocol; '
            'not treated as a difficult patient redo.'
        ),
        'Why marked': r['why'],
    } for _, r in calibration_df.iterrows()])
else:
    calibration_tracker = pd.DataFrame()

by_rater_c = (
    tracker_df.groupby(['Assigned to', 'Email', 'Status'], dropna=False)
    .size().reset_index(name='Number of tasks')
    .sort_values(['Assigned to', 'Status'])
    if len(tracker_df) else pd.DataFrame(columns=['Assigned to', 'Email', 'Status', 'Number of tasks'])
)
by_flag = (
    tracker_df.groupby('Issue type').size().reset_index(name='Number of tasks')
    if len(tracker_df) else pd.DataFrame(columns=['Issue type', 'Number of tasks'])
)
unique_cases_c = (
    tracker_df.groupby(['Patient ID', 'Side', 'Protocol'], as_index=False)
    .agg(**{
        'People assigned': ('Assigned to', 'nunique'),
        'Assigned names': ('Assigned to', lambda s: ', '.join(sorted(s.unique()))),
        'Max % away from cluster': ('% away from cluster', 'max'),
    })
    .sort_values(['Protocol', 'Patient ID', 'Side'])
    if len(tracker_df) else pd.DataFrame()
)

print('\n=== Consensus redo (cluster outsiders) ===')
if len(tracker_df):
    display(tracker_df.head(20))
    print('\n=== Workload by person ===')
    display(by_rater_c)
    print(
        f"\n{len(tracker_df)} redo tasks · {tracker_df['Assigned to'].nunique()} people · "
        f"{len(unique_cases_c)} unique case/protocol rows"
    )
else:
    print('No cluster-outsider redo tasks.')

print(f'\nDifficult cases: {len(difficult_tracker)}')
print(f'Calibration cases: {len(calibration_tracker)}')
if len(calib_summary_df):
    print('\n=== Protocol style summary ===')
    display(calib_summary_df)

# Persist CSVs
csv_path_c = out_dir / 'redo_queue_consensus.csv'
if len(tracker_df):
    tracker_df.to_csv(csv_path_c, index=False)
else:
    pd.DataFrame(columns=[
        'Task ID', 'Status', 'Assigned to', 'Patient ID', 'Side', 'Protocol'
    ]).to_csv(csv_path_c, index=False)
if len(redo_df):
    redo_df.to_csv(out_dir / 'redo_queue_consensus_details.csv', index=False)
difficult_tracker.to_csv(out_dir / 'redo_queue_difficult.csv', index=False)
calibration_tracker.to_csv(out_dir / 'redo_queue_calibration.csv', index=False)
calib_summary_df.to_csv(out_dir / 'redo_queue_calibration_summary.csv', index=False)
outcomes_df.to_csv(out_dir / 'redo_queue_consensus_outcomes.csv', index=False)

# Load §5.1 pair tracker
pair_tracker = pair_by_rater = pair_unique = pair_details = None
try:
    pair_csv = out_dir / 'redo_queue.csv'
    if pair_csv.exists():
        pair_tracker = pd.read_csv(pair_csv)
    pbr = out_dir / 'redo_queue_by_rater.csv'
    puc = out_dir / 'redo_queue_unique_cases.csv'
    pdet = out_dir / 'redo_queue_details.csv'
    if pbr.exists():
        pair_by_rater = pd.read_csv(pbr)
    if puc.exists():
        pair_unique = pd.read_csv(puc)
    if pdet.exists():
        pair_details = pd.read_csv(pdet)
except Exception as e:
    print(f'⚠ Could not load §5.1 pair tracker CSVs: {e}')

workload_parts = []
if pair_tracker is not None and len(pair_tracker):
    tmp = pair_tracker.groupby(['Assigned to', 'Email', 'Status'], dropna=False).size().reset_index(name='Number of tasks')
    tmp.insert(0, 'List', 'Pair disagreement')
    workload_parts.append(tmp)
if len(tracker_df):
    tmp = tracker_df.groupby(['Assigned to', 'Email', 'Status'], dropna=False).size().reset_index(name='Number of tasks')
    tmp.insert(0, 'List', 'Consensus redo')
    workload_parts.append(tmp)
# Difficult / Calibration are case lists (not per-person tasks) — show counts in summary rows
if len(difficult_tracker):
    workload_parts.append(pd.DataFrame([{
        'List': 'Difficult cases',
        'Assigned to': '(group review)',
        'Email': '',
        'Status': 'Open',
        'Number of tasks': len(difficult_tracker),
    }]))
if len(calibration_tracker):
    workload_parts.append(pd.DataFrame([{
        'List': 'Calibration cases',
        'Assigned to': '(landmark training)',
        'Email': '',
        'Status': 'Open',
        'Number of tasks': len(calibration_tracker),
    }]))
workload_all = (
    pd.concat(workload_parts, ignore_index=True).sort_values(['List', 'Assigned to', 'Status'])
    if workload_parts else pd.DataFrame()
)

xlsx_path = out_dir / 'redo_tracker.xlsx'
try:
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        _write_instructions_sheet(writer, 'Combined redo tracker (§5.1 + §5.2)', [
            'One workbook — pair disagreement plus cluster consensus:',
            '',
            'PAGES FOR RATERS',
            '  • Pair disagreement — unusual gap vs another rater (§5.1: ≥10% and outside BA band).',
            '  • Consensus redo — you are outside an agreeing cluster (≥2–3 raters within 5%; you ≥10% away).',
            '  • Workload by person — how many redo tasks you have.',
            '',
            'PAGES FOR MENTORS / GROUP (not personal redo)',
            '  • Difficult cases — no clear agreeing cluster (or split camps). Tricky patient; no assignee.',
            '  • Calibration cases — disagreement matches known high/low rater style camps for that protocol.',
            '  • Calibration summary — who tends high vs low per protocol.',
            '',
            'HOW TO USE',
            '1. Filter "Assigned to" to your name on Pair disagreement and Consensus redo.',
            '2. Re-annotate that Patient ID + Side + Protocol; set Status → Done.',
            '3. Do not treat Difficult / Calibration rows as personal redo queues.',
            '',
            'Details / outcomes CSVs are optional. Re-running the notebook overwrites this file.',
        ])

        if pair_tracker is not None and len(pair_tracker):
            pair_tracker.to_excel(writer, sheet_name='Pair disagreement', index=False)
            if pair_by_rater is not None:
                pair_by_rater.to_excel(writer, sheet_name='Pair — workload', index=False)
            if pair_unique is not None:
                pair_unique.to_excel(writer, sheet_name='Pair — unique cases', index=False)
            if pair_details is not None:
                pair_details.to_excel(writer, sheet_name='Pair — details', index=False)

        if len(tracker_df):
            tracker_df.to_excel(writer, sheet_name='Consensus redo', index=False)
            by_rater_c.to_excel(writer, sheet_name='Consensus — workload', index=False)
            by_flag.to_excel(writer, sheet_name='Consensus — issue types', index=False)
            if len(unique_cases_c):
                unique_cases_c.to_excel(writer, sheet_name='Consensus — unique cases', index=False)
            redo_df.to_excel(writer, sheet_name='Consensus — details', index=False)
        else:
            pd.DataFrame({'Note': ['No cluster-outsider redo tasks under current thresholds.']}).to_excel(
                writer, sheet_name='Consensus redo', index=False
            )

        if len(difficult_tracker):
            difficult_tracker.to_excel(writer, sheet_name='Difficult cases', index=False)
        else:
            pd.DataFrame({'Note': ['No difficult (no-cluster / split) cases.']}).to_excel(
                writer, sheet_name='Difficult cases', index=False
            )

        if len(calibration_tracker):
            calibration_tracker.to_excel(writer, sheet_name='Calibration cases', index=False)
        if len(calib_summary_df):
            calib_summary_df.to_excel(writer, sheet_name='Calibration summary', index=False)

        if len(workload_all):
            workload_all.to_excel(writer, sheet_name='Workload by person', index=False)

        if 'Pair disagreement' in writer.sheets:
            _style_tracker_sheet(
                writer.sheets['Pair disagreement'],
                status_header='Status',
                highlight_header='Priority focus (you look like the outlier)?',
                highlight_value='Yes — start here',
            )
        if 'Consensus redo' in writer.sheets and len(tracker_df):
            _style_tracker_sheet(writer.sheets['Consensus redo'], status_header='Status')
            from openpyxl.styles import PatternFill
            ws = writer.sheets['Consensus redo']
            headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
            status_col = headers.get('Status')
            fill = PatternFill('solid', fgColor='DDEBF7')
            for r in range(2, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    if status_col != c:
                        ws.cell(r, c).fill = fill
        if 'Difficult cases' in writer.sheets and len(difficult_tracker):
            _style_tracker_sheet(writer.sheets['Difficult cases'], status_header='Status')
        if 'Calibration cases' in writer.sheets and len(calibration_tracker):
            _style_tracker_sheet(writer.sheets['Calibration cases'], status_header='Status')
        if 'Workload by person' in writer.sheets:
            _style_tracker_sheet(writer.sheets['Workload by person'], status_header='Status')

    print(f'\nSaved combined tracker: {xlsx_path}')
except ImportError:
    print('\n⚠ openpyxl not installed — wrote CSV only.')

print(f'Saved: {csv_path_c}')
for legacy in ('redo_queue.xlsx', 'redo_queue_consensus.xlsx'):
    legacy_path = out_dir / legacy
    if legacy_path.exists():
        try:
            legacy_path.unlink()
            print(f'Removed legacy file: {legacy_path.name}')
        except Exception:
            pass
print(
    '\nOpen redo_tracker.xlsx — use "Pair disagreement" and "Consensus redo" for personal tasks; '
    '"Difficult cases" / "Calibration cases" for group review.'
)

"""
---
## Part III — Planar Landmark Agreement

Clinical-value agreement (Part II) tells us whether the *final number* is reproducible,
but a protocol can produce a stable angle from two raters who clicked in noticeably
different places if their errors happen to cancel out geometrically. This part asks the
finer-grained question directly: how consistent are the raw **(x, y) landmark clicks**
themselves, in the imaging plane?

**Scope note — planar only.** Every landmark in `stepPointsMmJson` also carries a `z`
(slice index) in `pointsJson`, but Z is not analyzed anywhere in this notebook. A rater's
slice choice is accepted as-is; only the in-plane (x, y) position is scored. This is a
deliberate simplification, not an oversight — it isolates "did they click the same spot
on the slice they chose" from "did they choose the same slice," which is a separate
question this version does not address.

**Key implementation facts (verified against real export):**
- **v4:** all `step_measurement` rows are endpoint-standardized and sulcus condyle
  lines are anatomically relabeled (Section 1.5) *before* any planar statistics.
- Merge key: `patientId + laterality + true_protocol + stepId` — not just `stepId`,
  because `condyle-line` appears in both `tt-tg` and `patellar-tilt`, and `lateral-line` /
  `medial-line` appear in both `sulcus-angle` and `sulcus-angle-3cm`.
- `pointCount = 3` for `distance` type: pt[0], pt[1] are the annotated endpoints;
  pt[2] is the foot-of-perpendicular auto-computed by the app (~0.2 mm from pt[1]).
  **Only pt[0] and pt[1] are used for planar error.**
- `pointCount = 2` for `perpendicular` type (trochlear-groove, tibial-tubercle):
  pt[0] = anchor on reference line, pt[1] = landmark. Both points included.
- `pointCount = 1` for `joint-line` (sagittal point, sulcus-angle-3cm only): a single
  point has no pairing partner for a planar-distance calculation and is excluded here.
- Coordinates in `stepPointsMmJson` are in **physical mm**, within the slice plane.
- With N raters, every step is compared across all $\\binom{N}{2}$ pairs, then pooled.

### 6. Planar Coordinate Variability and ICC

Two complementary views of the same question:
1. **Pairwise planar distance** (mm) between matched landmarks, per step — concrete and
   interpretable in millimetres, pooled across all rater pairs.
2. **Planar ICC** on the (x, y) magnitude — a single reliability coefficient per the
   usual ICC convention, computed on the balanced subset of landmarks every rater
   annotated.

Landmarks are matched between raters by the Hungarian algorithm (minimum-cost bipartite
matching on Euclidean XY distance) rather than by click order, since raters need not
click points in the same sequence. After Section 1.5 standardization, sulcus condyle
lines and line directions are already aligned — Hungarian matching remains useful for
perpendicular anchor points and any residual ordering ambiguity.
"""

import matplotlib.pyplot as plt
import seaborn as sns
from scipy.optimize import linear_sum_assignment


def parse_points(val):
    """Alias for Section 1.5 parser (used throughout Part III)."""
    return parse_points_json(val)


def effective_points(pts, point_count, meas_type):
    """Planar points actually used for comparison (drops the auto-computed foot-of-perpendicular)."""
    return pts[:2] if (meas_type == 'distance' and len(pts) == 3) else pts


def match_points_xy(pts_a, pts_b):
    """
    Pair two point sets by minimizing total Euclidean XY distance (Hungarian algorithm),
    so comparisons are robust to raters clicking landmarks in a different order.
    """
    n = len(pts_a)
    if n == 0 or n != len(pts_b):
        return pts_b

    a_xy = np.array([[p['x'], p['y']] for p in pts_a])
    b_xy = np.array([[p['x'], p['y']] for p in pts_b])

    cost_matrix = np.linalg.norm(a_xy[:, np.newaxis, :] - b_xy[np.newaxis, :, :], axis=2)
    row_ind, col_ind = linear_sum_assignment(cost_matrix)
    return [pts_b[i] for i in col_ind]


# ── Build per-rater step_measurement frames ──────────────────────────────────
step_cols = ['patientId', 'laterality', 'true_protocol', 'stepId', 'stepLabel',
             'plane', 'sliceIndex', 'measurementType', 'pointCount',
             'stepPointsMmJson', 'pointsJson']
step_frames = {
    rk: dataframes[rk][dataframes[rk]['recordType'] == 'step_measurement'][step_cols].copy()
    for rk in rater_keys
}
STEP_MERGE_KEYS = ['patientId', 'laterality', 'true_protocol', 'stepId']


def planar_distance_for_pair(ra, rb):
    """Per-landmark planar (XY) distance between two raters, for every shared step."""
    merged = step_frames[ra].merge(step_frames[rb], on=STEP_MERGE_KEYS, suffixes=('_a', '_b'))
    rows = []
    for _, row in merged.iterrows():
        mtype, pc = row['measurementType_a'], row['pointCount_a']
        if pc == 1:
            continue  # single-point steps have no planar pairing partner

        pts_a = effective_points(parse_points(row['stepPointsMmJson_a']), pc, mtype)
        pts_b = effective_points(parse_points(row['stepPointsMmJson_b']), pc, mtype)
        if not pts_a or not pts_b or len(pts_a) != len(pts_b):
            continue

        pts_b_matched = match_points_xy(pts_a, pts_b)
        for i, (p_a, p_b) in enumerate(zip(pts_a, pts_b_matched)):
            dist = np.linalg.norm([p_a['x'] - p_b['x'], p_a['y'] - p_b['y']])
            rows.append({
                'rater_a': ra, 'rater_b': rb,
                'patientId': row['patientId'], 'true_protocol': row['true_protocol'],
                'stepId': row['stepId'], 'stepLabel': row['stepLabel_a'],
                'pointIndex': i, 'dist_mm': dist,
            })
    return pd.DataFrame(rows)


pairwise_planar_dfs = {(ra, rb): planar_distance_for_pair(ra, rb) for ra, rb in rater_pairs}
planar_variability_df = (
    pd.concat(pairwise_planar_dfs.values(), ignore_index=True)
    if pairwise_planar_dfs else pd.DataFrame()
)

print("Planar coordinate variability (Hungarian-matched landmarks, all rater pairs pooled):")
if not planar_variability_df.empty:
    display(planar_variability_df.groupby('stepId')['dist_mm'].agg(['mean', 'std', 'max']).round(2))
    plt.figure(figsize=(14, 6))
    sns.barplot(data=planar_variability_df, x='stepId', y='dist_mm', errorbar='sd')
    plt.xticks(rotation=45, ha='right')
    plt.title('Planar Coordinate Variability by Step (Hungarian-Matched)')
    plt.ylabel('Planar Distance Error (mm)')
    plt.tight_layout()
    plt.show()
else:
    print('No overlapping data.')

# ── Planar ICC on the balanced (all-raters-present) subset ─────────────────
planar_icc_records = []
for rk in rater_keys:
    df = step_frames[rk]
    for _, row in df.iterrows():
        pts = effective_points(parse_points(row['stepPointsMmJson']), row['pointCount'], row['measurementType'])
        if not pts:
            continue
        for i, pt in enumerate(pts):
            planar_icc_records.append({
                'subject': f"{row['patientId']}_{row['laterality']}_{row['true_protocol']}_{row['stepId']}_pt{i}",
                'rater': rk.split(' | ')[0],
                'pos_magnitude': np.sqrt(pt['x'] ** 2 + pt['y'] ** 2),
            })

planar_long_df = pd.DataFrame(planar_icc_records)
# Collapse duplicate runs within the same rater/subject to one value (median) before pivoting.
planar_long_df = planar_long_df.groupby(['subject', 'rater'], as_index=False)['pos_magnitude'].median()

# Balanced subset: landmarks annotated by every rater.
planar_pivot_df = planar_long_df.pivot(index='subject', columns='rater', values='pos_magnitude').dropna()

if not planar_pivot_df.empty:
    print(f"\nPlanar ICC on a balanced subset of {len(planar_pivot_df)} landmark points "
          f"(all {len(rater_keys)} raters present):")

    balanced_long = planar_pivot_df.reset_index().melt(id_vars='subject', value_name='pos_magnitude')

    try:
        planar_icc_results = pg.intraclass_corr(
            data=balanced_long, targets='subject', raters='rater', ratings='pos_magnitude'
        )
        display(planar_icc_results)

        # Pingouin labels absolute agreement as 'ICC2' / 'ICC(A,1)' depending on version.
        agreement_row = planar_icc_results[
            planar_icc_results['Type'].str.contains(r'ICC\(A,1\)|ICC2', regex=True)
        ]
        consistency_row = planar_icc_results[
            planar_icc_results['Type'].str.contains(r'ICC\(C,1\)|ICC3', regex=True)
        ]

        if not agreement_row.empty:
            print(f"\nOverall Planar ICC (Absolute Agreement): {agreement_row['ICC'].values[0]:.3f}")
        if not consistency_row.empty:
            print(f"Overall Planar ICC (Consistency):        {consistency_row['ICC'].values[0]:.3f}")
        if agreement_row.empty and consistency_row.empty:
            print("Could not isolate a specific ICC type from the results table above.")
        else:
            print("\nInterpretation: agreement on (x, y) landmark position alone, by construction "
                  "blind to which slice (Z) the rater chose.")
    except Exception as e:
        print(f"Error calculating planar ICC on the balanced subset: {e}")
else:
    print("\nNo landmarks were found that all raters annotated — cannot compute a balanced planar ICC.")

"""
### 7. Outlier Triage — Largest Planar Disagreements

Surfaces the specific landmarks with the largest planar error, so they can be reviewed
manually (mislabeled laterality, wrong anatomical structure, genuine ambiguity, etc.).
This is purely descriptive — it does not feed back into the ICC or summary statistics
above.
"""

PLANAR_OUTLIER_THRESHOLD_MM = 100

extreme_outliers = planar_variability_df[
    planar_variability_df['dist_mm'] > PLANAR_OUTLIER_THRESHOLD_MM
].copy()

if not extreme_outliers.empty:
    print(f"Found {len(extreme_outliers)} landmarks with planar error > {PLANAR_OUTLIER_THRESHOLD_MM} mm.")
    outlier_summary = extreme_outliers.sort_values('dist_mm', ascending=False).head(20)
    display(outlier_summary)

    plt.figure(figsize=(10, 5))
    sns.stripplot(data=extreme_outliers, x='stepId', y='dist_mm', hue='stepId', legend=False)
    plt.xticks(rotation=45, ha='right')
    plt.title(f'Outlier Landmarks (planar error > {PLANAR_OUTLIER_THRESHOLD_MM} mm)')
    plt.ylabel('Planar Distance Error (mm)')
    plt.tight_layout()
    plt.show()
else:
    print(f"No extreme outliers (> {PLANAR_OUTLIER_THRESHOLD_MM} mm) found in the current matching set.")

"""
---
## Part IV — Ground Truth and Readiness

### 8. ML Ground Truth — Averaged Clinical Values

Averages **all raters'** clinical measurements (row-wise mean, ignoring missing raters)
into a single ground-truth label per case, weighted by that protocol's ICC from Part II.
This label is the clinical scalar (angle/ratio/mm) the protocol ultimately produces —
no per-landmark coordinates or Z information are involved, so it carries over unchanged
from a Z-aware version of this notebook.
"""

if comparison_df.empty:
    print("comparison_df is empty — run Section 3 first.")
else:
    ml_data = comparison_df.copy()

    # Calculate the mean ground truth across all available raters for each record
    ml_data['ground_truth'] = ml_data[value_cols].mean(axis=1, skipna=True)
    ml_data['ground_truth_std'] = ml_data[value_cols].std(axis=1, skipna=True)

    # Map ICC weights if available, default to 1.0 if not
    if 'icc_summary' in locals() and not icc_summary.empty:
        icc_map = icc_summary.set_index('Protocol')['ICC'].to_dict()
        ml_data['gt_icc_weight'] = ml_data['true_protocol'].map(icc_map).fillna(1.0)
    else:
        ml_data['gt_icc_weight'] = 1.0

    output_file = os.path.join(project_path, 'ml_ready_annotations.csv')
    ml_data.to_csv(output_file, index=False)

    print(f"--- ML Ground Truth Generated ---")
    print(f"Total unique measurement events: {len(ml_data)}")
    display(ml_data[['patientId', 'laterality', 'true_protocol', 'n_raters_present',
                      'ground_truth', 'ground_truth_std', 'gt_icc_weight']].head(10))

    print(f"\nSaved labels to: {output_file}")
    print(f"Average raters per record: {ml_data['n_raters_present'].mean():.2f}")

"""
### 9. Annotation Versioning and Duplicate-Run Checks

The README flags a production gap: neither CSV export includes `annotationVersion`,
`correctionOf`, or `modelVersion`. This section detects missing columns and flags
potential unversioned duplicate annotations, across all raters.
"""

RECOMMENDED_COLS = ['annotationVersion', 'correctionOf', 'modelVersion']
versioning_report = []
for rk in rater_keys:
    df = dataframes[rk]
    missing = [c for c in RECOMMENDED_COLS if c not in df.columns]
    present = [c for c in RECOMMENDED_COLS if c in df.columns]
    versioning_report.append({
        'Rater': rk,
        'Present': ', '.join(present) or 'none',
        'Missing': ', '.join(missing),
    })
display(pd.DataFrame(versioning_report))
if any(r['Missing'] for r in versioning_report):
    print("\n⚠  Add to sessionAnnotationCsv.ts and protocolMeasurementCsv.ts:")
    print("   annotationVersion : int, monotonically increasing")
    print("   correctionOf      : UUID of previous annotationId")
    print("   modelVersion      : model version that pre-populated this annotation")

# Duplicate detection: same patientId + laterality + true_protocol within ONE rater
print("\nChecking for duplicate protocol runs (same patient + laterality + true_protocol, within rater)...")
for rk in rater_keys:
    res = dataframes[rk][dataframes[rk]['recordType'] == 'protocol_result']
    dups = res.duplicated(subset=['patientId', 'laterality', 'true_protocol'], keep=False)
    n = dups.sum()
    print(f"  {rk}: {n} duplicate runs" + (" ← may be revised annotations" if n else " ✓"))
    if n:
        display(res[dups][['patientId', 'laterality', 'true_protocol', 'groupId', 'resultValue']]
                .sort_values('true_protocol').head(8))

# Timestamp check
print("\nTimestamp ranges:")
for rk in rater_keys:
    df = dataframes[rk]
    if 'timestamp' in df.columns:
        ts = pd.to_datetime(df['timestamp'], errors='coerce')
        print(f"  {rk}: {ts.min()} → {ts.max()}  ({ts.isna().sum()} unparseable)")
    else:
        print(f"  {rk}: no timestamp column")

# ── Self-consistency: foot-of-perpendicular check (per rater) ────────────────
# pt[2] for 3-point 'distance' steps is auto-computed by the app and should sit
# almost exactly on pt[1]; large deviations indicate an export inconsistency.
redundancy_checks = []
for rk in rater_keys:
    step = dataframes[rk][dataframes[rk]['recordType'] == 'step_measurement']
    for _, row in step[step['pointCount'] == 3].iterrows():
        pts = parse_points(row['stepPointsMmJson'])
        if len(pts) == 3:
            d12 = np.linalg.norm([pts[1]['x'] - pts[2]['x'], pts[1]['y'] - pts[2]['y']])
            redundancy_checks.append({
                'rater': rk, 'stepId': row['stepId'], 'dist_pt1_pt2_mm': d12
            })

redundancy_df = pd.DataFrame(redundancy_checks)
print("\nFoot-of-perpendicular check (expect dist ≈ 0), all raters:")
display(redundancy_df.groupby(['stepId'])['dist_pt1_pt2_mm'].describe().round(4))
large = redundancy_df[redundancy_df['dist_pt1_pt2_mm'] > 1.0]
if len(large):
    print(f"\n⚠ {len(large)} rows with pt[1]–pt[2] > 1 mm — possible export inconsistency:")
    display(large)
else:
    print("\n✓ All pt[1]–pt[2] distances ≤ 1 mm, across all raters.")

"""
### 10. Per-Protocol ML Readiness Summary

Counts per `true_protocol` (from `groupId`), summed across **all raters** — the only way
to correctly separate `sulcus-angle` from `sulcus-angle-3cm`, and to credit every
annotator's contribution regardless of how many files they exported.
"""

PROTOCOL_DEFS = {
    'tt-tg':            {'plane': 'axial',    'n_landmarks': 3, 'min_studies': 300},
    'insall-salvati':   {'plane': 'sagittal', 'n_landmarks': 2, 'min_studies': 300},
    'patellar-tilt':    {'plane': 'axial',    'n_landmarks': 2, 'min_studies': 300},
    'sulcus-angle':     {'plane': 'axial',    'n_landmarks': 2, 'min_studies': 300},
    'sulcus-angle-3cm': {'plane': 'axial',    'n_landmarks': 3, 'min_studies': 500,
                         'note': 'cross-plane; joint-line on sagittal + 2 lines on axial'},
    'caton-deschamps':  {'plane': 'sagittal', 'n_landmarks': 2, 'min_studies': 500,
                         'note': 'hardest; articular cartilage boundary'},
}

readiness_rows = []
for rk in rater_keys:
    df   = dataframes[rk]
    res  = df[df['recordType'] == 'protocol_result']
    step = df[df['recordType'] == 'step_measurement']
    for proto_id, info in PROTOCOL_DEFS.items():
        grp      = res[res['true_protocol'] == proto_id]
        step_grp = step[step['true_protocol'] == proto_id]
        n_studies = grp['patientId'].nunique()
        n_sides   = len(grp)
        n_slices  = step_grp['sliceIndex'].nunique() if 'sliceIndex' in step_grp.columns else 0
        status    = ('✓ Feasibility' if n_studies >= 100 else
                     '⚠ Pre-feasibility' if n_studies >= 50 else
                     '✗ Insufficient')
        readiness_rows.append({
            'Protocol':         proto_id,
            'Plane':            info['plane'],
            'Rater':            rk,
            'Unique patients':  n_studies,
            'Sides annotated':  n_sides,
            'Unique slices':    n_slices,
            'Unique steps':     step_grp['stepId'].nunique(),
            'Min target':       info['min_studies'],
            'Status':           status,
            'Notes':            info.get('note', ''),
        })

readiness_df = pd.DataFrame(readiness_rows)
print(f"Per-Protocol ML Readiness across {len(rater_keys)} rater(s) (true_protocol from groupId):")
display(readiness_df)

readiness_agg = (
    readiness_df.groupby('Protocol')['Unique patients']
    .sum().reset_index()
    .rename(columns={'Unique patients': 'Total patients (all raters)'})
)
readiness_agg['Status'] = readiness_agg['Total patients (all raters)'].apply(
    lambda n: '✓ Feasibility' if n >= 100 else ('⚠ Pre-feasibility' if n >= 50 else '✗ Insufficient')
)
print(f"\nCombined across all {len(rater_keys)} rater(s):")
display(readiness_agg)
